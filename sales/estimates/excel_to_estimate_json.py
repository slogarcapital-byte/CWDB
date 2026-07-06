"""
CWDB Excel quote -> estimate JSON reader

Reads the Quote Input sheet of a CWDB_Deck_Estimator workbook copy back into
the canonical estimate-JSON schema, so Excel-only quotes can feed
generate_work_order_pdf.py / generate_sow_pdf.py (and regenerate the PDF
estimate) without manual transcription.

Reuses the same single sources of truth the Streamlit app writes with:
  - QI cell map (sales/estimating/build_estimator_workbook.py)
  - INPUT_CELL_MAP (sales/estimating/streamlit_app/populate_workbook.py)
  - build_estimate_json engine (sales/estimates/deck_calculator.py)

The workbook stores only name + address for the client; pass phone/email on
the command line (they are required on the work order / contract).

Usage:
    python excel_to_estimate_json.py ../estimates/2026-06-03-overbeck-stain.xlsx \
        --phone "(715) 393-7145" --email doverbeck1@gmail.com \
        [--out _data/2026-06-03-overbeck-stain.json]
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent          # sales/estimates
SALES_DIR = SCRIPT_DIR.parent                          # sales/
ESTIMATING_DIR = SALES_DIR / 'estimating'
STREAMLIT_DIR = ESTIMATING_DIR / 'streamlit_app'

for p in (SCRIPT_DIR, ESTIMATING_DIR, STREAMLIT_DIR):
    if str(p) not in sys.path:
        sys.path.insert(0, str(p))

from populate_workbook import INPUT_CELL_MAP            # noqa: E402
from deck_calculator import build_estimate_json, load_pricing  # noqa: E402

# Fields that belong to client_info rather than engine inputs.
CLIENT_KEYS = {'client_name', 'address', 'estimator', 'quote_date'}

# Engine inputs that must be numeric; blank cells fall back to 0.
NUMERIC_DEFAULT_ZERO = {
    'length', 'depth', 'railing_lf', 'fascia_lf', 'stair_runs', 'stair_treads',
    'stair_landings', 'stain_sf', 'stain_coats', 'board_repairs',
    'joist_repair_lf', 'skirting_sf', 'lighting_fix', 'bench_count',
    'privacy_screen_lf',
}


def read_quote_input(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb['Quote Input']
    raw = {key: ws[cell].value for key, cell in INPUT_CELL_MAP.items()}

    inputs = {}
    for key, value in raw.items():
        if key in CLIENT_KEYS:
            continue
        if value is None and key in NUMERIC_DEFAULT_ZERO:
            value = 0
        if isinstance(value, float) and value == int(value):
            value = int(value)
        inputs[key] = value

    missing = [k for k, v in inputs.items() if v is None]
    if missing:
        raise SystemExit(
            f'Quote Input sheet is missing required values: {missing} '
            f'(workbook: {xlsx_path})')

    quote_date = raw.get('quote_date')
    if isinstance(quote_date, datetime):
        quote_date = quote_date.date()
    elif isinstance(quote_date, str):
        try:
            quote_date = date.fromisoformat(quote_date.strip())
        except ValueError:
            quote_date = None

    client = {
        'name': raw.get('client_name') or '',
        'address_line': raw.get('address') or '',
    }
    return inputs, client, quote_date


def main():
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument('xlsx_path', help='estimator workbook copy (.xlsx)')
    p.add_argument('--phone', default='',
                   help='client phone (not stored in the workbook)')
    p.add_argument('--email', default='',
                   help='client email (not stored in the workbook)')
    p.add_argument('--out', default=None,
                   help='output JSON path; defaults to _data/<xlsx stem>.json')
    p.add_argument('--force', action='store_true',
                   help='overwrite an existing JSON')
    args = p.parse_args()

    xlsx_path = Path(args.xlsx_path).resolve()
    inputs, client, quote_date = read_quote_input(xlsx_path)
    client['phone'] = args.phone
    client['email'] = args.email
    if not args.phone or not args.email:
        print('WARNING: client phone/email blank; required on the work order '
              'and contract. Pass --phone/--email.', file=sys.stderr)

    db = load_pricing()
    estimate = build_estimate_json(inputs, client, db, today=quote_date)

    out = (Path(args.out).resolve() if args.out
           else SCRIPT_DIR / '_data' / f'{xlsx_path.stem}.json')
    if out.exists() and not args.force:
        raise SystemExit(f'{out} already exists; pass --force to overwrite.')
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(estimate, indent=2, ensure_ascii=False) + '\n',
                   encoding='utf-8')
    total = estimate['_meta']['computed_sell_price']
    print(f'JSON written: {out}')
    print(f"Estimate #{estimate['estimate_number']}  ·  "
          f"{estimate['client']['name']}  ·  ${total:,}")


if __name__ == '__main__':
    main()
