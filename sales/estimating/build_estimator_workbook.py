"""
CWDB Deck Estimator v1 - Workbook Generator
Central Wisconsin Deck Builders, LLC

Builds CWDB_Deck_Estimator_v1.xlsx from pricing-db.json.

Layout:
  Sheet 1: Quote Input       -- user-facing input form + live cost calculations
  Sheet 2: Pricing DB        -- editable pricing assumptions
  Sheet 3: Project Type Behavior -- read-only matrix that drives the engine
  Sheet 4: Estimate Summary  -- three-tier comparison (where applicable)
  Sheet 5: Client Range      -- proposal-ready text block
  Sheet 6: Scope Builder     -- boilerplate scope language per project type
  Sheet 7: Notes             -- usage instructions + as-of timestamp

Run:
  python build_estimator_workbook.py

Output:
  ./CWDB_Deck_Estimator_v1.xlsx (next to this script)
"""

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                             NamedStyle)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.workbook.defined_name import DefinedName

# ---------------------------------------------------------------------------
# CWDB brand
# ---------------------------------------------------------------------------
ORANGE = "E54C00"
SLATE = "323434"
GREY = "646760"
SKY = "83B2CF"
LIGHT_BG = "F7F4F1"
LIGHT_GREY = "EFEFEF"
INPUT_YELLOW = "FFF8DC"
CALC_GREY = "F4F4F4"
NA_GREY = "DDDDDD"

THIN = Side(border_style="thin", color=GREY)
THICK = Side(border_style="medium", color=ORANGE)
HAIR = Side(border_style="hair", color=GREY)

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
def fill(color):
    return PatternFill("solid", fgColor=color)


def title_font():
    return Font(name="Calibri", size=18, bold=True, color=ORANGE)


def big_money_font():
    return Font(name="Calibri", size=22, bold=True, color=ORANGE)


def section_font():
    return Font(name="Calibri", size=12, bold=True, color="FFFFFF")


def label_font(bold=False):
    return Font(name="Calibri", size=10, bold=bold, color=SLATE)


def calc_font():
    return Font(name="Calibri", size=10, color=GREY, italic=True)


def result_font():
    return Font(name="Calibri", size=11, bold=True, color=SLATE)


CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
RIGHT = Alignment(horizontal="right", vertical="center")
WRAP_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def section_header(ws, row, text, span=3, color=ORANGE):
    """Write a section header row with orange fill, white text."""
    ws.cell(row=row, column=1).value = text
    ws.cell(row=row, column=1).font = section_font()
    ws.cell(row=row, column=1).fill = fill(color)
    ws.cell(row=row, column=1).alignment = LEFT
    for c in range(2, span + 1):
        ws.cell(row=row, column=c).fill = fill(color)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 22


def section_header_at(ws, row, start_col, end_col, text, color=ORANGE):
    """Section header at a specific column range."""
    cell = ws.cell(row=row, column=start_col)
    cell.value = text
    cell.font = section_font()
    cell.fill = fill(color)
    cell.alignment = LEFT
    for c in range(start_col + 1, end_col + 1):
        ws.cell(row=row, column=c).fill = fill(color)
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row,
                   end_column=end_col)
    ws.row_dimensions[row].height = 22


def label_cell(ws, row, col, text, bold=False):
    c = ws.cell(row=row, column=col)
    c.value = text
    c.font = label_font(bold=bold)
    c.alignment = LEFT


def input_cell(ws, row, col, default=None, number_format=None):
    c = ws.cell(row=row, column=col)
    if default is not None:
        c.value = default
    c.font = label_font()
    c.fill = fill(INPUT_YELLOW)
    c.alignment = LEFT
    c.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    if number_format:
        c.number_format = number_format
    return c


def calc_cell(ws, row, col, formula, number_format="#,##0.00"):
    c = ws.cell(row=row, column=col)
    c.value = formula
    c.font = calc_font()
    c.fill = fill(CALC_GREY)
    c.alignment = RIGHT
    c.number_format = number_format
    c.border = Border(left=HAIR, right=HAIR, top=HAIR, bottom=HAIR)
    return c


def db_cell(ws, row, col, value, number_format=None, editable=True):
    c = ws.cell(row=row, column=col)
    c.value = value
    c.font = label_font()
    c.alignment = LEFT
    if editable:
        c.fill = fill(INPUT_YELLOW)
    else:
        c.fill = fill(LIGHT_GREY)
    c.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    if number_format:
        c.number_format = number_format
    return c


def header_cell(ws, row, col, text):
    c = ws.cell(row=row, column=col)
    c.value = text
    c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    c.fill = fill(SLATE)
    c.alignment = CENTER
    c.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    return c


def add_dv(ws, range_str, options=None, formula=None):
    """Add a data validation dropdown."""
    if options is not None:
        # quote each option, wrap in commas
        formula_str = '"' + ",".join(options) + '"'
        dv = DataValidation(type="list", formula1=formula_str, allow_blank=True)
    else:
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Pick from the list"
    dv.errorTitle = "Invalid selection"
    ws.add_data_validation(dv)
    dv.add(range_str)


# ---------------------------------------------------------------------------
# Load pricing DB
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent
PRICING_DB_PATH = SCRIPT_DIR / "pricing-db.json"
OUTPUT_PATH = SCRIPT_DIR / "CWDB_Deck_Estimator_v1.xlsx"


def load_pricing():
    with open(PRICING_DB_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


# ---------------------------------------------------------------------------
# Shared layout maps. Populated by build_pricing_db / build_project_behavior
# (called first in main) and read by the formula-building sheets so no formula
# hard-codes a row that shifts when the data grows. Every value is a full,
# absolute, sheet-qualified A1 reference string.
# ---------------------------------------------------------------------------
PDB = {}   # Pricing DB ranges/cells, keyed by semantic name
PB = {}    # Project Type Behavior ranges, keyed by column letter ("A".."J")


def _pdb_col(col, start, end):
    return f"'Pricing DB'!${col}${start}:${col}${end}"


def _pdb_cell(col, row):
    return f"'Pricing DB'!${col}${row}"


# ---------------------------------------------------------------------------
# Build Pricing DB sheet (cursor-based; collision-proof; records PDB ranges)
# ---------------------------------------------------------------------------
def build_pricing_db_v2(wb, db):
    ws = wb.create_sheet("Pricing DB")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 48
    ws.column_dimensions["F"].width = 38

    ws["A1"] = "CWDB Pricing Database"
    ws["A1"].font = title_font()
    ws.merge_cells("A1:F1")
    ws["A2"] = (f"As-of {db['as_of']} · Edit yellow cells. Regenerate workbook "
                f"to reflect changes. Review quarterly.")
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=GREY)
    ws.merge_cells("A2:F2")

    cur = 5  # running row cursor

    def headers(row, labels):
        for col, text in enumerate(labels, start=1):
            header_cell(ws, row, col, text)

    # ---- Decking ----------------------------------------------------------
    section_header(ws, cur, "Decking Materials", span=6)
    headers(cur + 1, ["Name", "Sell $/SF", "Material $/SF", "Waste %",
                      "Description", "Source"])
    s = cur + 2
    for i, m in enumerate(db["decking_materials"]):
        r = s + i
        db_cell(ws, r, 1, m["name"], editable=False)
        db_cell(ws, r, 2, m["sell_per_sf"], number_format='"$"#,##0.00')
        db_cell(ws, r, 3, m["material_per_sf"], number_format='"$"#,##0.00')
        db_cell(ws, r, 4, m["waste_pct"], number_format='0.0%')
        db_cell(ws, r, 5, m["description"], editable=False)
        db_cell(ws, r, 6, m["source"], editable=False)
    e = s + len(db["decking_materials"]) - 1
    PDB["dk_name"] = _pdb_col("A", s, e)
    PDB["dk_sell"] = _pdb_col("B", s, e)
    PDB["dk_mat"] = _pdb_col("C", s, e)
    PDB["dk_waste"] = _pdb_col("D", s, e)
    cur = e + 2

    # ---- Railing ----------------------------------------------------------
    section_header(ws, cur, "Railing Materials", span=6)
    headers(cur + 1, ["Name", "Sell $/LF", "Material $/LF", "Description",
                      "Source", ""])
    s = cur + 2
    for i, m in enumerate(db["railing_materials"]):
        r = s + i
        db_cell(ws, r, 1, m["name"], editable=False)
        db_cell(ws, r, 2, m["sell_per_lf"], number_format='"$"#,##0')
        db_cell(ws, r, 3, m["material_per_lf"], number_format='"$"#,##0')
        db_cell(ws, r, 4, m["description"], editable=False)
        db_cell(ws, r, 5, m["source"], editable=False)
    e = s + len(db["railing_materials"]) - 1
    PDB["rl_name"] = _pdb_col("A", s, e)
    PDB["rl_sell"] = _pdb_col("B", s, e)
    cur = e + 2

    # ---- Framing ----------------------------------------------------------
    section_header(ws, cur, "Framing Materials", span=6)
    headers(cur + 1, ["Name", "Sell $/SF", "Material $/SF", "Description",
                      "Source", ""])
    s = cur + 2
    for i, m in enumerate(db["framing_materials"]):
        r = s + i
        db_cell(ws, r, 1, m["name"], editable=False)
        db_cell(ws, r, 2, m["sell_per_sf"], number_format='"$"#,##0.00')
        db_cell(ws, r, 3, m["material_per_sf"], number_format='"$"#,##0.00')
        db_cell(ws, r, 4, m["description"], editable=False)
        db_cell(ws, r, 5, m["source"], editable=False)
    e = s + len(db["framing_materials"]) - 1
    PDB["fr_name"] = _pdb_col("A", s, e)
    PDB["fr_sell"] = _pdb_col("B", s, e)
    cur = e + 2

    # ---- Demo Rates -------------------------------------------------------
    section_header(ws, cur, "Demo Rates by Project Type", span=6)
    headers(cur + 1, ["Project Type", "Demo $/SF", "", "", "", ""])
    s = cur + 2
    for i, pt in enumerate(db["project_types"]):
        r = s + i
        db_cell(ws, r, 1, pt["name"], editable=False)
        db_cell(ws, r, 2, db["demo_rates_per_sf"].get(pt["key"], 0),
                number_format='"$"#,##0.00')
    e = s + len(db["project_types"]) - 1
    PDB["demo_name"] = _pdb_col("A", s, e)
    PDB["demo_rate"] = _pdb_col("B", s, e)
    cur = e + 2

    # ---- Stain Rates ------------------------------------------------------
    section_header(ws, cur, "Stain Rates", span=6)
    headers(cur + 1, ["Stain Type", "Coats", "$/SF (combined)", "Key", "", ""])
    s = cur + 2
    for i, sr in enumerate(db["stain_rates_per_sf"]):
        r = s + i
        db_cell(ws, r, 1, sr["type"], editable=False)
        db_cell(ws, r, 2, sr["coats"], number_format='0')
        db_cell(ws, r, 3, sr["per_sf"], number_format='"$"#,##0.00')
        db_cell(ws, r, 4, f"{sr['type']}|{sr['coats']}", editable=False)
    e = s + len(db["stain_rates_per_sf"]) - 1
    PDB["stain_rate"] = _pdb_col("C", s, e)
    PDB["stain_key"] = _pdb_col("D", s, e)
    cur = e + 2

    # ---- Repair Rates -----------------------------------------------------
    section_header(ws, cur, "Repair Rates", span=6)
    rr = db["repair_rates"]
    repair_rows = [
        ("Board Replacement $/board", rr["board_replacement_per_board"], "repair_board"),
        ("Joist Repair $/LF", rr["joist_repair_per_lf"], "repair_joist"),
        ("Hardware Allowance $/job", rr["hardware_allowance"], "repair_hardware"),
        ("Inspection-only Allowance $/job", rr["inspection_only_allowance"], "repair_inspection"),
    ]
    s = cur + 1
    for i, (label, value, key) in enumerate(repair_rows):
        r = s + i
        db_cell(ws, r, 1, label, editable=False)
        db_cell(ws, r, 2, value, number_format='"$"#,##0')
        PDB[key] = _pdb_cell("B", r)
    cur = s + len(repair_rows) + 1

    # ---- Condition Multipliers -------------------------------------------
    cm = db["condition_multipliers"]
    for title, key, prefix in [
        ("Condition Multipliers - Height", "height", "mh"),
        ("Condition Multipliers - Grade", "grade", "mg"),
        ("Condition Multipliers - Complexity", "complexity", "mc"),
        ("Condition Multipliers - Market Load", "market_load", "mm"),
    ]:
        section_header(ws, cur, title, span=6)
        s = cur + 1
        for i, m in enumerate(cm[key]):
            r = s + i
            db_cell(ws, r, 1, m["value"], editable=False)
            db_cell(ws, r, 2, m["multiplier"], number_format='0.00"x"')
        e = s + len(cm[key]) - 1
        PDB[f"{prefix}_name"] = _pdb_col("A", s, e)
        PDB[f"{prefix}_mult"] = _pdb_col("B", s, e)
        cur = e + 2

    # ---- Stair Pricing ----------------------------------------------------
    section_header(ws, cur, "Stair Pricing", span=6)
    sp = db["stair_pricing"]
    stair_rows = [
        ("Stair Base Setup", sp["base_setup"], "stair_base"),
        ("Cost per Tread", sp["per_tread"], "stair_tread"),
        ("Cost per Landing", sp["per_landing"], "stair_landing"),
        ("Wraparound Stair Premium", sp["wraparound_premium"], "stair_wrap"),
        ("Double Border Stairs $/tread", sp["double_border_per_tread"], "stair_dblborder"),
    ]
    s = cur + 1
    for i, (label, value, key) in enumerate(stair_rows):
        r = s + i
        db_cell(ws, r, 1, label, editable=False)
        db_cell(ws, r, 2, value, number_format='"$"#,##0')
        PDB[key] = _pdb_cell("B", r)
    cur = s + len(stair_rows) + 1

    # ---- Border Pricing ---------------------------------------------------
    section_header(ws, cur, "Border Pricing", span=6)
    db_cell(ws, cur + 1, 1, "Pencil Border $/SF", editable=False)
    db_cell(ws, cur + 1, 2, db["border_pricing"]["pencil_per_sf"], number_format='"$"#,##0.00')
    db_cell(ws, cur + 2, 1, "Double Border $/SF", editable=False)
    db_cell(ws, cur + 2, 2, db["border_pricing"]["double_per_sf"], number_format='"$"#,##0.00')
    PDB["border_double"] = _pdb_cell("B", cur + 2)
    cur = cur + 4

    # ---- Other Per-Unit ---------------------------------------------------
    section_header(ws, cur, "Other Per-Unit Rates", span=6)
    db_cell(ws, cur + 1, 1, "Fascia $/LF", editable=False)
    db_cell(ws, cur + 1, 2, db["fascia_per_lf"], number_format='"$"#,##0')
    PDB["fascia"] = _pdb_cell("B", cur + 1)
    db_cell(ws, cur + 2, 1, "Skirting / Privacy Wall $/SF", editable=False)
    db_cell(ws, cur + 2, 2, db["skirting_per_sf"], number_format='"$"#,##0')
    PDB["skirting"] = _pdb_cell("B", cur + 2)
    cur = cur + 4

    # ---- Adders -----------------------------------------------------------
    section_header(ws, cur, "Adders (Optional Upgrades)", span=6)
    headers(cur + 1, ["Adder", "Unit", "Price", "Key", "", ""])
    s = cur + 2
    for i, ad in enumerate(db["adders"]):
        r = s + i
        db_cell(ws, r, 1, ad["name"], editable=False)
        db_cell(ws, r, 2, ad["unit"], editable=False)
        db_cell(ws, r, 3, ad["price"], number_format='"$"#,##0')
        db_cell(ws, r, 4, ad["key"], editable=False)
    e = s + len(db["adders"]) - 1
    PDB["adder_price"] = _pdb_col("C", s, e)
    PDB["adder_key"] = _pdb_col("D", s, e)
    cur = e + 2

    # ---- Allowances (Defaults) -------------------------------------------
    section_header(ws, cur, "Allowances (Defaults)", span=6)
    alw = db["allowances"]
    alw_rows = [
        ("Permit / Engineering", alw["permit_engineering_default"], "alw_permit_def"),
        ("Dumpster / Cleanup", alw["dumpster_cleanup_default"], "alw_dump_def"),
        ("Mobilization Minimum", alw["mobilization_minimum_default"], "alw_mob_def"),
        ("Default Misc", alw["misc_default"], "alw_misc_def"),
    ]
    s = cur + 1
    for i, (label, value, key) in enumerate(alw_rows):
        r = s + i
        db_cell(ws, r, 1, label, editable=False)
        db_cell(ws, r, 2, value, number_format='"$"#,##0')
        PDB[key] = _pdb_cell("B", r)
    cur = s + len(alw_rows) + 1

    # ---- Margin & Contingency --------------------------------------------
    section_header(ws, cur, "Margin & Contingency", span=6)
    mc = db["margin_and_contingency"]
    db_cell(ws, cur + 1, 1, "Default Target Gross Margin", editable=False)
    db_cell(ws, cur + 1, 2, mc["default_margin"], number_format='0.0%')
    db_cell(ws, cur + 2, 1, "Default Contingency Range (+/-)", editable=False)
    db_cell(ws, cur + 2, 2, mc["default_contingency"], number_format='0.0%')
    cur = cur + 4

    # ---- Allowances by Project Type --------------------------------------
    section_header(ws, cur, "Allowances by Project Type (auto-defaults)", span=6)
    headers(cur + 1, ["Project Type", "Permit $", "Dumpster $", "Mobilization $", "", ""])
    s = cur + 2
    for i, a in enumerate(db["allowances_by_project_type"]):
        r = s + i
        db_cell(ws, r, 1, a["name"], editable=False)
        db_cell(ws, r, 2, a["permit"], number_format='"$"#,##0')
        db_cell(ws, r, 3, a["dumpster"], number_format='"$"#,##0')
        db_cell(ws, r, 4, a["mobilization"], number_format='"$"#,##0')
    e = s + len(db["allowances_by_project_type"]) - 1
    PDB["alwbt_name"] = _pdb_col("A", s, e)
    PDB["alwbt_permit"] = _pdb_col("B", s, e)
    PDB["alwbt_dump"] = _pdb_col("C", s, e)
    PDB["alwbt_mob"] = _pdb_col("D", s, e)
    cur = e + 2

    # ---- Fence Materials (rate per LF by height) -------------------------
    section_header(ws, cur, "Fence Materials ($/LF by height)", span=6)
    fence_heights = db.get("fence_heights", ["4", "5", "6", "8"])
    headers(cur + 1, ["Material"] + [f"{h} ft" for h in fence_heights] + [""])
    # numeric height header row used for column MATCH
    hdr_row = cur + 1
    for j, h in enumerate(fence_heights):
        ws.cell(row=hdr_row, column=2 + j).value = int(h)
    s = cur + 2
    for i, fm in enumerate(db.get("fence_materials", [])):
        r = s + i
        db_cell(ws, r, 1, fm["name"], editable=False)
        for j, h in enumerate(fence_heights):
            rate = fm["cost_per_lf_by_height"].get(h)
            db_cell(ws, r, 2 + j, rate if rate is not None else 0,
                    number_format='"$"#,##0.00')
    n_fence = len(db.get("fence_materials", []))
    e = s + max(n_fence, 1) - 1
    last_col = get_column_letter(1 + len(fence_heights))
    PDB["fence_name"] = _pdb_col("A", s, e)
    PDB["fence_hdr"] = f"'Pricing DB'!$B${hdr_row}:${last_col}${hdr_row}"
    PDB["fence_rates"] = f"'Pricing DB'!$B${s}:${last_col}${e}"
    cur = e + 2

    # ---- Fence Gates / Other ---------------------------------------------
    section_header(ws, cur, "Fence Gates & Other", span=6)
    gates = {g["key"]: g for g in db.get("fence_gates", [])}
    rows = [
        ("Walk gate $/each", gates.get("walk", {}).get("cost_each", 0), "fence_walk"),
        ("Drive / double gate $/each", gates.get("drive", {}).get("cost_each", 0), "fence_drive"),
        ("Tear-out existing fence $/LF", db.get("fence_tearout_per_lf", 0), "fence_tearout"),
        ("Fence mobilization default", db.get("fence_mobilization_default", 0), "fence_mobil"),
    ]
    s = cur + 1
    for i, (label, value, key) in enumerate(rows):
        r = s + i
        db_cell(ws, r, 1, label, editable=False)
        db_cell(ws, r, 2, value, number_format='"$"#,##0')
        PDB[key] = _pdb_cell("B", r)
    cur = s + len(rows) + 1

    # ---- Materials Unit Costs --------------------------------------------
    section_header(ws, cur, "Materials Unit Costs (edit yellow cells)", span=6)
    headers(cur + 1, ["Key (do not edit)", "Description", "Unit", "Price",
                      "Source", ""])
    s = cur + 2
    for i, item in enumerate(db["materials_unit_costs"]):
        r = s + i
        db_cell(ws, r, 1, item["key"], editable=False)
        db_cell(ws, r, 2, item["description"], editable=False)
        db_cell(ws, r, 3, item["unit"], editable=False)
        db_cell(ws, r, 4, item["price"], number_format='"$"#,##0.00')
        db_cell(ws, r, 5, item["source"], editable=False)
    e = s + len(db["materials_unit_costs"]) - 1
    PDB["muc_key"] = _pdb_col("A", s, e)
    PDB["muc_price"] = _pdb_col("D", s, e)
    cur = e + 3

    # ---- Metadata ---------------------------------------------------------
    section_header(ws, cur, "Metadata", span=6)
    db_cell(ws, cur + 1, 1, "Pricing As-Of", editable=False)
    db_cell(ws, cur + 1, 2, db["as_of"], editable=False)
    db_cell(ws, cur + 2, 1, "Schema Version", editable=False)
    db_cell(ws, cur + 2, 2, db["schema_version"], editable=False)

    return ws


# ---------------------------------------------------------------------------
# Build Pricing DB sheet
# ---------------------------------------------------------------------------
def build_pricing_db(wb, db):
    """
    Pricing DB sheet layout (cell addresses are stable, formulas reference
    them by absolute address from the Quote Input sheet).

    Row anchors (DO NOT MOVE without updating Quote Input formulas):
      A5  : 'Decking Materials' section header
      A6  : decking data header row (Name | Sell$/SF | Mat$/SF | Waste | Desc | Source)
      A7-A11 : 5 decking rows

      A14 : 'Railing Materials' header
      A15 : railing data header row
      A16-A20 : 5 railing rows

      A23 : 'Framing Materials' header
      A24 : framing data header row
      A25-A26 : 2 framing rows

      A29 : 'Demo Rates by Project Type'
      A30 : header
      A31-A35 : 5 rows

      A38 : 'Stain Rates'
      A39 : header
      A40-A45 : 6 rows  (3 types x 2 coats)

      A48 : 'Repair Rates'
      A49-A52 : 4 rows

      A55 : 'Condition Multipliers - Height'
      A56-A57 : 2 rows
      A60 : 'Grade'
      A61-A63 : 3 rows
      A66 : 'Complexity'
      A67-A69 : 3 rows
      A72 : 'Market Load'
      A73-A75 : 3 rows

      A78 : 'Stair Pricing'
      A79-A83 : 5 rows (key, value)

      A86 : 'Border Pricing'
      A87-A88 : 2 rows

      A91 : 'Other Per-Unit Rates'
      A92 : Fascia $/LF
      A93 : Skirting $/SF

      A96 : 'Adders'
      A97 : header
      A98-A102 : 5 rows

      A105 : 'Allowances (Defaults)'
      A106-A109 : 4 rows

      A112 : 'Margin & Contingency'
      A113-A114 : 2 rows

      A117 : 'Metadata'
      A118-A119 : 2 rows
    """
    ws = wb.create_sheet("Pricing DB")

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 48
    ws.column_dimensions["F"].width = 38

    ws["A1"] = "CWDB Pricing Database"
    ws["A1"].font = title_font()
    ws.merge_cells("A1:F1")

    ws["A2"] = (f"As-of {db['as_of']} · Edit yellow cells. "
                f"Regenerate workbook to reflect changes. Review quarterly.")
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=GREY)
    ws.merge_cells("A2:F2")

    # ---- Decking Materials -------------------------------------------------
    section_header(ws, 5, "Decking Materials", span=6)
    for col, text in enumerate(["Name", "Sell $/SF", "Material $/SF",
                                "Waste %", "Description", "Source"], start=1):
        header_cell(ws, 6, col, text)

    for i, mat in enumerate(db["decking_materials"]):
        r = 7 + i
        db_cell(ws, r, 1, mat["name"], editable=False)
        db_cell(ws, r, 2, mat["sell_per_sf"], number_format='"$"#,##0.00')
        db_cell(ws, r, 3, mat["material_per_sf"], number_format='"$"#,##0.00')
        db_cell(ws, r, 4, mat["waste_pct"], number_format='0.0%')
        db_cell(ws, r, 5, mat["description"], editable=False)
        db_cell(ws, r, 6, mat["source"], editable=False)

    # ---- Railing Materials -------------------------------------------------
    section_header(ws, 14, "Railing Materials", span=6)
    for col, text in enumerate(["Name", "Sell $/LF", "Material $/LF",
                                "Description", "Source", ""], start=1):
        header_cell(ws, 15, col, text)

    for i, mat in enumerate(db["railing_materials"]):
        r = 16 + i
        db_cell(ws, r, 1, mat["name"], editable=False)
        db_cell(ws, r, 2, mat["sell_per_lf"], number_format='"$"#,##0')
        db_cell(ws, r, 3, mat["material_per_lf"], number_format='"$"#,##0')
        db_cell(ws, r, 4, mat["description"], editable=False)
        db_cell(ws, r, 5, mat["source"], editable=False)

    # ---- Framing Materials -------------------------------------------------
    section_header(ws, 23, "Framing Materials", span=6)
    for col, text in enumerate(["Name", "Sell $/SF", "Material $/SF",
                                "Description", "Source", ""], start=1):
        header_cell(ws, 24, col, text)

    for i, mat in enumerate(db["framing_materials"]):
        r = 25 + i
        db_cell(ws, r, 1, mat["name"], editable=False)
        db_cell(ws, r, 2, mat["sell_per_sf"], number_format='"$"#,##0.00')
        db_cell(ws, r, 3, mat["material_per_sf"], number_format='"$"#,##0.00')
        db_cell(ws, r, 4, mat["description"], editable=False)
        db_cell(ws, r, 5, mat["source"], editable=False)

    # ---- Demo Rates --------------------------------------------------------
    section_header(ws, 29, "Demo Rates by Project Type", span=6)
    header_cell(ws, 30, 1, "Project Type")
    header_cell(ws, 30, 2, "Demo $/SF")

    proj_types = db["project_types"]
    for i, pt in enumerate(proj_types):
        r = 31 + i
        rate = db["demo_rates_per_sf"].get(pt["key"], 0)
        db_cell(ws, r, 1, pt["name"], editable=False)
        db_cell(ws, r, 2, rate, number_format='"$"#,##0.00')

    # ---- Stain Rates -------------------------------------------------------
    section_header(ws, 38, "Stain Rates", span=6)
    header_cell(ws, 39, 1, "Stain Type")
    header_cell(ws, 39, 2, "Coats")
    header_cell(ws, 39, 3, "$/SF (combined)")
    header_cell(ws, 39, 4, "Key")  # Concatenated lookup key: "Type|Coats"
    for i, sr in enumerate(db["stain_rates_per_sf"]):
        r = 40 + i
        db_cell(ws, r, 1, sr["type"], editable=False)
        db_cell(ws, r, 2, sr["coats"], number_format='0')
        db_cell(ws, r, 3, sr["per_sf"], number_format='"$"#,##0.00')
        db_cell(ws, r, 4, f"{sr['type']}|{sr['coats']}", editable=False)

    # ---- Repair Rates ------------------------------------------------------
    section_header(ws, 47, "Repair Rates", span=6)
    rr = db["repair_rates"]
    repair_rows = [
        ("Board Replacement $/board", rr["board_replacement_per_board"]),
        ("Joist Repair $/LF", rr["joist_repair_per_lf"]),
        ("Hardware Allowance $/job", rr["hardware_allowance"]),
        ("Inspection-only Allowance $/job", rr["inspection_only_allowance"]),
    ]
    for i, (label, value) in enumerate(repair_rows):
        r = 48 + i
        db_cell(ws, r, 1, label, editable=False)
        db_cell(ws, r, 2, value, number_format='"$"#,##0')

    # ---- Multipliers -------------------------------------------------------
    cm = db["condition_multipliers"]

    section_header(ws, 53, "Condition Multipliers - Height", span=6)
    for i, m in enumerate(cm["height"]):
        r = 54 + i
        db_cell(ws, r, 1, m["value"], editable=False)
        db_cell(ws, r, 2, m["multiplier"], number_format='0.00"x"')

    section_header(ws, 57, "Condition Multipliers - Grade", span=6)
    for i, m in enumerate(cm["grade"]):
        r = 58 + i
        db_cell(ws, r, 1, m["value"], editable=False)
        db_cell(ws, r, 2, m["multiplier"], number_format='0.00"x"')

    section_header(ws, 62, "Condition Multipliers - Complexity", span=6)
    for i, m in enumerate(cm["complexity"]):
        r = 63 + i
        db_cell(ws, r, 1, m["value"], editable=False)
        db_cell(ws, r, 2, m["multiplier"], number_format='0.00"x"')

    section_header(ws, 67, "Condition Multipliers - Market Load", span=6)
    for i, m in enumerate(cm["market_load"]):
        r = 68 + i
        db_cell(ws, r, 1, m["value"], editable=False)
        db_cell(ws, r, 2, m["multiplier"], number_format='0.00"x"')

    # ---- Stair Pricing -----------------------------------------------------
    section_header(ws, 72, "Stair Pricing", span=6)
    sp = db["stair_pricing"]
    stair_rows = [
        ("Stair Base Setup", sp["base_setup"]),
        ("Cost per Tread", sp["per_tread"]),
        ("Cost per Landing", sp["per_landing"]),
        ("Wraparound Stair Premium", sp["wraparound_premium"]),
        ("Double Border Stairs $/tread", sp["double_border_per_tread"]),
    ]
    for i, (label, value) in enumerate(stair_rows):
        r = 73 + i
        db_cell(ws, r, 1, label, editable=False)
        db_cell(ws, r, 2, value, number_format='"$"#,##0')

    # ---- Border Pricing ----------------------------------------------------
    section_header(ws, 79, "Border Pricing", span=6)
    border_rows = [
        ("Pencil Border $/SF", db["border_pricing"]["pencil_per_sf"]),
        ("Double Border $/SF", db["border_pricing"]["double_per_sf"]),
    ]
    for i, (label, value) in enumerate(border_rows):
        r = 80 + i
        db_cell(ws, r, 1, label, editable=False)
        db_cell(ws, r, 2, value, number_format='"$"#,##0.00')

    # ---- Other Per-Unit ----------------------------------------------------
    section_header(ws, 83, "Other Per-Unit Rates", span=6)
    db_cell(ws, 84, 1, "Fascia $/LF", editable=False)
    db_cell(ws, 84, 2, db["fascia_per_lf"], number_format='"$"#,##0')
    db_cell(ws, 85, 1, "Skirting / Privacy Wall $/SF", editable=False)
    db_cell(ws, 85, 2, db["skirting_per_sf"], number_format='"$"#,##0')

    # ---- Adders ------------------------------------------------------------
    section_header(ws, 88, "Adders (Optional Upgrades)", span=6)
    for col, text in enumerate(["Adder", "Unit", "Price", "Key"], start=1):
        header_cell(ws, 89, col, text)

    for i, ad in enumerate(db["adders"]):
        r = 90 + i
        db_cell(ws, r, 1, ad["name"], editable=False)
        db_cell(ws, r, 2, ad["unit"], editable=False)
        db_cell(ws, r, 3, ad["price"], number_format='"$"#,##0')
        db_cell(ws, r, 4, ad["key"], editable=False)

    # ---- Allowances --------------------------------------------------------
    section_header(ws, 96, "Allowances (Defaults)", span=6)
    alw = db["allowances"]
    alw_rows = [
        ("Permit / Engineering", alw["permit_engineering_default"]),
        ("Dumpster / Cleanup", alw["dumpster_cleanup_default"]),
        ("Mobilization Minimum", alw["mobilization_minimum_default"]),
        ("Default Misc", alw["misc_default"]),
    ]
    for i, (label, value) in enumerate(alw_rows):
        r = 97 + i
        db_cell(ws, r, 1, label, editable=False)
        db_cell(ws, r, 2, value, number_format='"$"#,##0')

    # ---- Margin & Contingency ---------------------------------------------
    section_header(ws, 102, "Margin & Contingency", span=6)
    mc = db["margin_and_contingency"]
    db_cell(ws, 103, 1, "Default Target Gross Margin", editable=False)
    db_cell(ws, 103, 2, mc["default_margin"], number_format='0.0%')
    db_cell(ws, 104, 1, "Default Contingency Range (+/-)", editable=False)
    db_cell(ws, 104, 2, mc["default_contingency"], number_format='0.0%')

    # ---- Allowances by Project Type ---------------------------------------
    # IMPORTANT: rows 114-118 are referenced by Quote Input B58/B59/B60.
    # Do not move without updating those formulas.
    section_header(ws, 113, "Allowances by Project Type (auto-defaults)", span=6)
    for col, text in enumerate(["Project Type", "Permit $", "Dumpster $",
                                "Mobilization $", "", ""], start=1):
        header_cell(ws, 114, col, text)
    for i, alw in enumerate(db["allowances_by_project_type"]):
        r = 115 + i
        db_cell(ws, r, 1, alw["name"], editable=False)
        db_cell(ws, r, 2, alw["permit"], number_format='"$"#,##0')
        db_cell(ws, r, 3, alw["dumpster"], number_format='"$"#,##0')
        db_cell(ws, r, 4, alw["mobilization"], number_format='"$"#,##0')

    # ---- Materials Unit Costs ---------------------------------------------
    # IMPORTANT: rows 130-160 are referenced by Materials List sheet formulas
    # via MATCH against the key in column A.
    section_header(ws, 128, "Materials Unit Costs (edit yellow cells)", span=6)
    for col, text in enumerate(["Key (do not edit)", "Description", "Unit",
                                "Price", "Source", ""], start=1):
        header_cell(ws, 129, col, text)
    for i, item in enumerate(db["materials_unit_costs"]):
        r = 130 + i
        db_cell(ws, r, 1, item["key"], editable=False)
        db_cell(ws, r, 2, item["description"], editable=False)
        db_cell(ws, r, 3, item["unit"], editable=False)
        db_cell(ws, r, 4, item["price"], number_format='"$"#,##0.00')
        db_cell(ws, r, 5, item["source"], editable=False)

    # ---- Metadata ----------------------------------------------------------
    meta_start = 130 + len(db["materials_unit_costs"]) + 3
    section_header(ws, meta_start, "Metadata", span=6)
    db_cell(ws, meta_start + 1, 1, "Pricing As-Of", editable=False)
    db_cell(ws, meta_start + 1, 2, db["as_of"], editable=False)
    db_cell(ws, meta_start + 2, 1, "Schema Version", editable=False)
    db_cell(ws, meta_start + 2, 2, db["schema_version"], editable=False)
    db_cell(ws, meta_start + 3, 1, "Market", editable=False)
    db_cell(ws, meta_start + 3, 2, db["market"], editable=False)
    ws.cell(row=meta_start + 3, column=2).alignment = WRAP_LEFT
    ws.merge_cells(start_row=meta_start + 3, start_column=2,
                   end_row=meta_start + 3, end_column=5)

    return ws


# ---------------------------------------------------------------------------
# Build Project Type Behavior sheet
# ---------------------------------------------------------------------------
def build_project_behavior(wb, db):
    """
    Project Type Behavior matrix.
    Row anchors:
      A1 : title
      A3 : header row (Project Type | Demo | Frame | Deck | Rail | Stair
                       | Footing | Stain | Multipliers | Description)
      A4-A8 : 5 project types

    Engine formulas in Quote Input read INDEX/MATCH against B4:I8.
    """
    ws = wb.create_sheet("Project Type Behavior")

    ws.column_dimensions["A"].width = 36
    for col in "BCDEFGHI":
        ws.column_dimensions[col].width = 11
    ws.column_dimensions["J"].width = 60

    ws["A1"] = "Project Type Behavior Matrix"
    ws["A1"].font = title_font()
    ws.merge_cells("A1:J1")

    ws["A2"] = ("Read-only. Each project type tells the engine which cost "
                "buckets to include (Y), exclude (N), allow optionally (OPT), "
                "or inspect only. Multipliers: Y = all four, Light = Height "
                "+ Grade only, N = none.")
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=GREY)
    ws["A2"].alignment = WRAP_LEFT
    ws.merge_cells("A2:J2")
    ws.row_dimensions[2].height = 36

    headers = ["Project Type", "Demo", "Frame", "Deck", "Rail", "Stair",
               "Footing", "Stain", "Multipliers", "Description"]
    for i, h in enumerate(headers, start=1):
        header_cell(ws, 3, i, h)

    matrix_keys = ["demo", "frame", "deck", "rail", "stair", "footing",
                   "stain", "multipliers"]
    for i, pt in enumerate(db["project_types"]):
        r = 4 + i
        db_cell(ws, r, 1, pt["name"], editable=False)
        for j, key in enumerate(matrix_keys):
            c = db_cell(ws, r, 2 + j, pt["matrix"][key], editable=False)
            c.alignment = CENTER
        desc = db_cell(ws, r, 10, pt["description"], editable=False)
        desc.alignment = WRAP_LEFT
        ws.row_dimensions[r].height = 30

    n = len(db["project_types"])
    last = 3 + n
    PB["A"] = f"'Project Type Behavior'!$A$4:$A${last}"
    for col in "BCDEFGHI":
        PB[col] = f"'Project Type Behavior'!${col}$4:${col}${last}"
    PB["last_row"] = last

    return ws


# ---------------------------------------------------------------------------
# Build Quote Input sheet (the brain)
# ---------------------------------------------------------------------------
# Cell address constants. Quote Input sheet only.
class QI:
    # Inputs - column B
    CLIENT_NAME = "B7"
    ADDRESS = "B8"
    ESTIMATOR = "B9"
    QUOTE_DATE = "B10"

    PROJECT_TYPE = "B13"

    LENGTH = "B16"
    DEPTH = "B17"
    DECK_SF = "B18"

    DECKING_MAT = "B21"
    RAILING_MAT = "B22"
    FRAMING_MAT = "B23"

    HEIGHT_COND = "B26"
    GRADE_COND = "B27"
    COMPLEXITY_COND = "B28"
    MARKET_COND = "B29"

    BORDER_STYLE = "B32"
    RAILING_LF = "B33"
    FASCIA_LF = "B34"
    STAIR_RUNS = "B35"
    STAIR_TREADS = "B36"
    STAIR_LANDINGS = "B37"
    WRAPAROUND = "B38"

    STAIN_SF = "B41"
    STAIN_TYPE = "B42"
    STAIN_COATS = "B43"

    BOARD_REPAIRS = "B46"
    JOIST_REPAIR_LF = "B47"
    HARDWARE_INC = "B48"

    SKIRTING_SF = "B51"
    LIGHTING_FIX = "B52"
    BENCH_COUNT = "B53"
    PRIV_SCREEN_LF = "B54"
    HOT_TUB = "B55"

    PERMIT_ALW = "B58"
    DUMPSTER_ALW = "B59"
    MOBIL_ALW = "B60"
    MISC_ALW = "B61"

    TARGET_MARGIN = "B64"
    CONTINGENCY = "B65"

    # Cost components - column E
    FRAMING_PSF = "E7"
    DECKING_PSF = "E8"
    DECKING_WASTE = "E9"
    RAILING_PLF = "E10"

    HEIGHT_MULT = "E13"
    GRADE_MULT = "E14"
    COMPLEXITY_MULT = "E15"
    MARKET_MULT = "E16"
    COMBINED_MULT = "E17"

    BASE_PACKAGE = "E20"
    DEMO_COST = "E21"
    BORDER_COST = "E22"
    RAILING_COST = "E23"
    FASCIA_COST = "E24"
    STAIR_COST = "E25"
    STAIN_COST = "E26"
    REPAIR_COST = "E27"
    SKIRTING_COST = "E28"
    LIGHTING_COST = "E29"
    BENCH_COST = "E30"
    PRIV_SCREEN_COST = "E31"
    HOT_TUB_COST = "E32"
    PERMIT_COST = "E33"
    DUMPSTER_COST = "E34"
    MOBIL_COST = "E35"
    MISC_COST = "E36"
    SUBTOTAL = "E37"

    SELL_PRICE = "E40"
    LOW_RANGE = "E41"
    HIGH_RANGE = "E42"
    PRICE_PER_SF = "E43"

    # Color inputs (populated; no calc impact unless a color carries an override)
    DECKING_COLOR = "B79"
    RAILING_COLOR = "B80"
    STAIN_COLOR = "B81"

    # Fence inputs (left column, below the margin block)
    FENCE_MAT = "B72"
    FENCE_HEIGHT = "B73"
    FENCE_LF = "B74"
    WALK_GATES = "B75"
    DRIVE_GATES = "B76"
    TEAROUT_LF = "B77"
    FENCE_COLOR = "B78"

    # Fence calc (right column)
    FENCE_RATE = "E45"
    FENCE_RUN = "E46"
    FENCE_GATES_COST = "E47"
    FENCE_TEAROUT_COST = "E48"
    FENCE_ALW = "E49"
    FENCE_SUBTOTAL = "E50"


def matrix_lookup(field_col_letter):
    """Returns an Excel INDEX/MATCH formula fragment that looks up a value
    from the Project Type Behavior matrix for the currently-selected
    project type. field_col_letter is B-I (Demo-Multipliers)."""
    return (f"INDEX({PB[field_col_letter]}, MATCH({QI.PROJECT_TYPE}, "
            f"{PB['A']}, 0))")


def build_quote_input(wb, db):
    ws = wb.create_sheet("Quote Input", 0)  # First sheet

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 22

    # =====================================================================
    # ROW 1-5: TOP DASHBOARD - big sell price
    # =====================================================================
    ws["A1"] = "CWDB Deck Project Quote"
    ws["A1"].font = title_font()
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 28

    ws["A2"] = "Sell Price"
    ws["A2"].font = Font(name="Calibri", size=12, bold=True, color=GREY)
    ws["A2"].alignment = LEFT
    ws.merge_cells("A2:B2")
    ws["C2"] = f"={QI.SELL_PRICE}"
    ws["C2"].font = big_money_font()
    ws["C2"].alignment = LEFT
    ws["C2"].number_format = '"$"#,##0'
    ws.merge_cells("C2:F2")
    ws.row_dimensions[2].height = 34

    ws["A3"] = "Range (low - high)"
    ws["A3"].font = Font(name="Calibri", size=10, color=GREY)
    ws["A3"].alignment = LEFT
    ws.merge_cells("A3:B3")
    ws["C3"] = (f'=TEXT({QI.LOW_RANGE},"$#,##0")&" - "&'
                f'TEXT({QI.HIGH_RANGE},"$#,##0")')
    ws["C3"].font = result_font()
    ws["C3"].alignment = LEFT
    ws.merge_cells("C3:F3")

    ws["A4"] = "Sell Price / SF"
    ws["A4"].font = Font(name="Calibri", size=10, color=GREY)
    ws["A4"].alignment = LEFT
    ws.merge_cells("A4:B4")
    ws["C4"] = f"=IFERROR({QI.PRICE_PER_SF}, 0)"
    ws["C4"].font = result_font()
    ws["C4"].alignment = LEFT
    ws["C4"].number_format = '"$"#,##0.00'

    # =====================================================================
    # LEFT COLUMN (A-C): INPUTS  |  RIGHT COLUMN (D-F): COST COMPONENTS
    # =====================================================================

    # ---- Project Info (rows 6-10) -----------------------------------------
    section_header_at(ws, 6, 1, 3, "Project Info")
    label_cell(ws, 7, 1, "Client Name")
    input_cell(ws, 7, 2, "Sample Homeowner")
    label_cell(ws, 8, 1, "Project Address")
    input_cell(ws, 8, 2, "")
    label_cell(ws, 9, 1, "Estimator")
    input_cell(ws, 9, 2, "James Slogar")
    label_cell(ws, 10, 1, "Quote Date")
    input_cell(ws, 10, 2, db["as_of"])

    # ---- Cost Components header (rows 6-7) on right ----------------------
    section_header_at(ws, 6, 4, 6, "Cost Components")

    label_cell(ws, 7, 4, "Framing $/SF", bold=True)
    calc_cell(ws, 7, 5,
              f"=IFERROR(INDEX({PDB['fr_sell']}, MATCH({QI.FRAMING_MAT}, "
              f"{PDB['fr_name']}, 0)), 0)",
              number_format='"$"#,##0.00')

    label_cell(ws, 8, 4, "Decking $/SF", bold=True)
    calc_cell(ws, 8, 5,
              f"=IFERROR(INDEX({PDB['dk_sell']}, MATCH({QI.DECKING_MAT}, "
              f"{PDB['dk_name']}, 0)), 0)",
              number_format='"$"#,##0.00')

    label_cell(ws, 9, 4, "Decking Waste %")
    calc_cell(ws, 9, 5,
              f"=IFERROR(INDEX({PDB['dk_waste']}, MATCH({QI.DECKING_MAT}, "
              f"{PDB['dk_name']}, 0)), 0)",
              number_format='0.0%')

    label_cell(ws, 10, 4, "Railing $/LF", bold=True)
    calc_cell(ws, 10, 5,
              f"=IFERROR(INDEX({PDB['rl_sell']}, MATCH({QI.RAILING_MAT}, "
              f"{PDB['rl_name']}, 0)), 0)",
              number_format='"$"#,##0')

    # ---- Project Type (rows 12-13) ---------------------------------------
    # IMPORTANT: dropdown references the Project Type Behavior sheet's A4:A8
    # range (NOT an inline CSV) because project type names contain commas
    # which break Excel's inline-list parser.
    section_header_at(ws, 12, 1, 3, "Project Type")
    label_cell(ws, 13, 1, "Project Type")
    input_cell(ws, 13, 2, db["project_types"][4]["name"])  # Default: Full Tear-Out
    add_dv(ws, "B13", formula=PB['A'])

    # Multipliers
    section_header_at(ws, 12, 4, 6, "Multipliers")
    label_cell(ws, 13, 4, "Height multiplier")
    calc_cell(ws, 13, 5,
              f"=IFERROR(INDEX({PDB['mh_mult']}, MATCH({QI.HEIGHT_COND}, "
              f"{PDB['mh_name']}, 0)), 1)",
              number_format='0.00"x"')
    label_cell(ws, 14, 4, "Grade multiplier")
    calc_cell(ws, 14, 5,
              f"=IFERROR(INDEX({PDB['mg_mult']}, MATCH({QI.GRADE_COND}, "
              f"{PDB['mg_name']}, 0)), 1)",
              number_format='0.00"x"')
    label_cell(ws, 15, 4, "Complexity multiplier")
    calc_cell(ws, 15, 5,
              f"=IFERROR(INDEX({PDB['mc_mult']}, MATCH({QI.COMPLEXITY_COND}, "
              f"{PDB['mc_name']}, 0)), 1)",
              number_format='0.00"x"')
    label_cell(ws, 16, 4, "Market multiplier")
    calc_cell(ws, 16, 5,
              f"=IFERROR(INDEX({PDB['mm_mult']}, MATCH({QI.MARKET_COND}, "
              f"{PDB['mm_name']}, 0)), 1)",
              number_format='0.00"x"')

    # Combined multiplier - respects matrix.Multipliers (N / Light / Y)
    label_cell(ws, 17, 4, "Combined multiplier", bold=True)
    calc_cell(ws, 17, 5,
              f'=IF({matrix_lookup("I")}="N", 1, '
              f'IF({matrix_lookup("I")}="Light", '
              f'{QI.HEIGHT_MULT}*{QI.GRADE_MULT}, '
              f'{QI.HEIGHT_MULT}*{QI.GRADE_MULT}*'
              f'{QI.COMPLEXITY_MULT}*{QI.MARKET_MULT}))',
              number_format='0.000"x"')

    # ---- Dimensions (rows 15-18) ----------------------------------------
    section_header_at(ws, 15, 1, 3, "Dimensions")
    label_cell(ws, 16, 1, "Deck Length")
    input_cell(ws, 16, 2, 20, number_format="0")
    ws["C16"] = "feet"
    ws["C16"].font = label_font()
    label_cell(ws, 17, 1, "Deck Depth")
    input_cell(ws, 17, 2, 15, number_format="0")
    ws["C17"] = "feet"
    ws["C17"].font = label_font()
    label_cell(ws, 18, 1, "Deck SF (calc)")
    ws["B18"] = f"={QI.LENGTH}*{QI.DEPTH}"
    ws["B18"].font = label_font(bold=True)
    ws["B18"].fill = fill(CALC_GREY)
    ws["B18"].number_format = "0"
    ws["C18"] = "sq ft"
    ws["C18"].font = label_font()

    # ---- Cost Components: line items (rows 19-37) -----------------------
    section_header_at(ws, 19, 4, 6, "Line Items")

    # E20 - Base Package (framing + decking conditional on matrix)
    label_cell(ws, 20, 4, "Base Package")
    calc_cell(ws, 20, 5,
              f'={QI.DECK_SF}*'
              f'(IF({matrix_lookup("C")}="Y", {QI.FRAMING_PSF}, 0) + '
              f'IF({matrix_lookup("D")}="Y", {QI.DECKING_PSF}*(1+{QI.DECKING_WASTE}), 0))'
              f'*{QI.COMBINED_MULT}',
              number_format='"$"#,##0')

    # E21 - Demo (project-type-specific rate)
    label_cell(ws, 21, 4, "Demo")
    calc_cell(ws, 21, 5,
              f'=IF({matrix_lookup("B")}="N", 0, '
              f'{QI.DECK_SF}*IFERROR(INDEX({PDB["demo_rate"]}, '
              f'MATCH({QI.PROJECT_TYPE}, {PDB["demo_name"]}, 0)), 0))',
              number_format='"$"#,##0')

    # E22 - Border (only when deck="Y")
    label_cell(ws, 22, 4, "Border")
    calc_cell(ws, 22, 5,
              f'=IF({matrix_lookup("D")}<>"Y", 0, '
              f'IF({QI.BORDER_STYLE}="Double Border", '
              f'{QI.DECK_SF}*{PDB["border_double"]}, 0))',
              number_format='"$"#,##0')

    # E23 - Railing
    label_cell(ws, 23, 4, "Railing")
    calc_cell(ws, 23, 5,
              f'=IF(OR({matrix_lookup("E")}="N"), 0, '
              f'{QI.RAILING_LF}*{QI.RAILING_PLF})',
              number_format='"$"#,##0')

    # E24 - Fascia
    label_cell(ws, 24, 4, "Fascia")
    calc_cell(ws, 24, 5,
              f'=IF({matrix_lookup("D")}<>"Y", 0, '
              f'{QI.FASCIA_LF}*{PDB["fascia"]})',
              number_format='"$"#,##0')

    # E25 - Stairs
    label_cell(ws, 25, 4, "Stairs")
    calc_cell(ws, 25, 5,
              f'=IF(OR({matrix_lookup("F")}="N", {QI.STAIR_RUNS}=0), 0, '
              f"{PDB['stair_base']} + "
              f"{QI.STAIR_TREADS}*{PDB['stair_tread']} + "
              f"{QI.STAIR_LANDINGS}*{PDB['stair_landing']} + "
              f'IF({QI.WRAPAROUND}="Yes", {PDB["stair_wrap"]}, 0) + '
              f'IF({QI.BORDER_STYLE}="Double Border", '
              f"{QI.STAIR_TREADS}*{PDB['stair_dblborder']}, 0))",
              number_format='"$"#,##0')

    # E26 - Stain
    label_cell(ws, 26, 4, "Stain")
    calc_cell(ws, 26, 5,
              f'=IF({matrix_lookup("H")}<>"Y", 0, '
              f"{QI.STAIN_SF}*IFERROR(INDEX({PDB['stain_rate']}, "
              f"MATCH({QI.STAIN_TYPE}&\"|\"&{QI.STAIN_COATS}, "
              f"{PDB['stain_key']}, 0)), 0))",
              number_format='"$"#,##0')

    # E27 - Repairs
    label_cell(ws, 27, 4, "Repairs")
    calc_cell(ws, 27, 5,
              f'=IF(AND({QI.PROJECT_TYPE}<>"Stain + Minor Repairs", '
              f'{QI.PROJECT_TYPE}<>"Resurface (Boards Only)"), 0, '
              f"{QI.BOARD_REPAIRS}*{PDB['repair_board']} + "
              f"{QI.JOIST_REPAIR_LF}*{PDB['repair_joist']} + "
              f"IF({QI.HARDWARE_INC}=\"Yes\", {PDB['repair_hardware']}, 0) + "
              f'IF({QI.PROJECT_TYPE}="Resurface (Boards Only)", '
              f"{PDB['repair_inspection']}, 0))",
              number_format='"$"#,##0')

    # E28 - Skirting (allowed for Y deck only)
    label_cell(ws, 28, 4, "Skirting / Privacy Wall")
    calc_cell(ws, 28, 5,
              f'=IF({matrix_lookup("D")}<>"Y", 0, '
              f"{QI.SKIRTING_SF}*{PDB['skirting']})",
              number_format='"$"#,##0')

    # E29 - Lighting (allowed when deck or stain)
    label_cell(ws, 29, 4, "Lighting Fixtures")
    calc_cell(ws, 29, 5,
              f"={QI.LIGHTING_FIX}*IFERROR(INDEX({PDB['adder_price']}, "
              f"MATCH(\"lighting\", {PDB['adder_key']}, 0)), 0)",
              number_format='"$"#,##0')

    # E30 - Built-In Benches
    label_cell(ws, 30, 4, "Built-In Benches")
    calc_cell(ws, 30, 5,
              f"={QI.BENCH_COUNT}*IFERROR(INDEX({PDB['adder_price']}, "
              f"MATCH(\"bench\", {PDB['adder_key']}, 0)), 0)",
              number_format='"$"#,##0')

    # E31 - Privacy Screen
    label_cell(ws, 31, 4, "Privacy Screen")
    calc_cell(ws, 31, 5,
              f"={QI.PRIV_SCREEN_LF}*IFERROR(INDEX({PDB['adder_price']}, "
              f"MATCH(\"privacy_screen\", {PDB['adder_key']}, 0)), 0)",
              number_format='"$"#,##0')

    # E32 - Hot Tub
    label_cell(ws, 32, 4, "Hot Tub Structural")
    calc_cell(ws, 32, 5,
              f'=IF({QI.HOT_TUB}="Yes", '
              f"IFERROR(INDEX({PDB['adder_price']}, "
              f"MATCH(\"hot_tub\", {PDB['adder_key']}, 0)), 0), 0)",
              number_format='"$"#,##0')

    # E33-E36 - Allowances (echo from input cells)
    label_cell(ws, 33, 4, "Permit / Engineering")
    calc_cell(ws, 33, 5, f"={QI.PERMIT_ALW}", number_format='"$"#,##0')
    label_cell(ws, 34, 4, "Dumpster / Cleanup")
    calc_cell(ws, 34, 5, f"={QI.DUMPSTER_ALW}", number_format='"$"#,##0')
    label_cell(ws, 35, 4, "Mobilization")
    calc_cell(ws, 35, 5, f"={QI.MOBIL_ALW}", number_format='"$"#,##0')
    label_cell(ws, 36, 4, "Misc")
    calc_cell(ws, 36, 5, f"={QI.MISC_ALW}", number_format='"$"#,##0')

    # E37 - SUBTOTAL
    label_cell(ws, 37, 4, "SUBTOTAL (Cost)", bold=True)
    sub = calc_cell(ws, 37, 5,
                    f'=IF({QI.PROJECT_TYPE}="Fence", {QI.FENCE_SUBTOTAL}, '
                    f"SUM({QI.BASE_PACKAGE}:{QI.MISC_COST}))",
                    number_format='"$"#,##0')
    sub.font = Font(name="Calibri", size=11, bold=True, color=SLATE)
    sub.fill = fill(LIGHT_BG)

    # ---- Final pricing (rows 39-43) --------------------------------------
    section_header_at(ws, 39, 4, 6, "Final Price")
    label_cell(ws, 40, 4, "Sell Price", bold=True)
    sell = calc_cell(ws, 40, 5,
                     f"={QI.SUBTOTAL}/(1-{QI.TARGET_MARGIN})",
                     number_format='"$"#,##0')
    sell.font = Font(name="Calibri", size=12, bold=True, color=ORANGE)
    sell.fill = fill(LIGHT_BG)

    label_cell(ws, 41, 4, "Low Range")
    calc_cell(ws, 41, 5,
              f"={QI.SELL_PRICE}*(1-{QI.CONTINGENCY})",
              number_format='"$"#,##0')
    label_cell(ws, 42, 4, "High Range")
    calc_cell(ws, 42, 5,
              f"={QI.SELL_PRICE}*(1+{QI.CONTINGENCY})",
              number_format='"$"#,##0')
    label_cell(ws, 43, 4, "Sell Price / unit")
    calc_cell(ws, 43, 5,
              f'=IF({QI.PROJECT_TYPE}="Fence", '
              f"IF({QI.FENCE_LF}=0, 0, {QI.SELL_PRICE}/{QI.FENCE_LF}), "
              f"IF({QI.DECK_SF}=0, 0, {QI.SELL_PRICE}/{QI.DECK_SF}))",
              number_format='"$"#,##0.00')

    # ---- Fence calc (right column, rows 44-50) ---------------------------
    section_header_at(ws, 44, 4, 6, "Fence Calc (used when Project Type = Fence)")
    label_cell(ws, 45, 4, "Fence rate $/LF")
    calc_cell(ws, 45, 5,
              f"=IFERROR(INDEX({PDB['fence_rates']}, "
              f"MATCH({QI.FENCE_MAT}, {PDB['fence_name']}, 0), "
              f"MATCH(VALUE({QI.FENCE_HEIGHT}), {PDB['fence_hdr']}, 0)), 0)",
              number_format='"$"#,##0.00')
    label_cell(ws, 46, 4, "Fence run")
    calc_cell(ws, 46, 5, f"={QI.FENCE_LF}*{QI.FENCE_RATE}", number_format='"$"#,##0')
    label_cell(ws, 47, 4, "Gates")
    calc_cell(ws, 47, 5,
              f"={QI.WALK_GATES}*{PDB['fence_walk']}+{QI.DRIVE_GATES}*{PDB['fence_drive']}",
              number_format='"$"#,##0')
    label_cell(ws, 48, 4, "Tear-out")
    calc_cell(ws, 48, 5, f"={QI.TEAROUT_LF}*{PDB['fence_tearout']}", number_format='"$"#,##0')
    label_cell(ws, 49, 4, "Fence allowances")
    calc_cell(ws, 49, 5,
              f"={QI.PERMIT_ALW}+{QI.DUMPSTER_ALW}+{QI.MOBIL_ALW}+{QI.MISC_ALW}",
              number_format='"$"#,##0')
    label_cell(ws, 50, 4, "Fence subtotal")
    calc_cell(ws, 50, 5,
              f"={QI.FENCE_RUN}+{QI.FENCE_GATES_COST}+{QI.FENCE_TEAROUT_COST}+{QI.FENCE_ALW}",
              number_format='"$"#,##0')

    # =====================================================================
    # LEFT COLUMN CONTINUED: more inputs
    # =====================================================================

    # ---- Materials (rows 20-23) ------------------------------------------
    section_header_at(ws, 20, 1, 3, "Materials")
    label_cell(ws, 21, 1, "Decking Material")
    input_cell(ws, 21, 2, db["decking_materials"][3]["name"])  # Trex Select default
    add_dv(ws, "B21", options=[m["name"] for m in db["decking_materials"]])

    label_cell(ws, 22, 1, "Railing Material")
    input_cell(ws, 22, 2, db["railing_materials"][2]["name"])  # Trex Aluminum default
    add_dv(ws, "B22", options=[m["name"] for m in db["railing_materials"]])

    label_cell(ws, 23, 1, "Framing Material")
    input_cell(ws, 23, 2, db["framing_materials"][0]["name"])  # KDAT default
    add_dv(ws, "B23", options=[m["name"] for m in db["framing_materials"]])

    # ---- Site Conditions (rows 25-29) ------------------------------------
    section_header_at(ws, 25, 1, 3, "Site Conditions")
    cm = db["condition_multipliers"]

    label_cell(ws, 26, 1, "Height")
    input_cell(ws, 26, 2, cm["height"][0]["value"])
    add_dv(ws, "B26", options=[m["value"] for m in cm["height"]])

    label_cell(ws, 27, 1, "Grade")
    input_cell(ws, 27, 2, cm["grade"][0]["value"])
    add_dv(ws, "B27", options=[m["value"] for m in cm["grade"]])

    label_cell(ws, 28, 1, "Complexity")
    input_cell(ws, 28, 2, cm["complexity"][0]["value"])
    add_dv(ws, "B28", options=[m["value"] for m in cm["complexity"]])

    label_cell(ws, 29, 1, "Market Load")
    input_cell(ws, 29, 2, cm["market_load"][0]["value"])
    add_dv(ws, "B29", options=[m["value"] for m in cm["market_load"]])

    # ---- Scope Detail (rows 31-38) ---------------------------------------
    section_header_at(ws, 31, 1, 3, "Scope Detail")

    label_cell(ws, 32, 1, "Border Style")
    input_cell(ws, 32, 2, "Pencil Border")
    add_dv(ws, "B32", options=db["border_options"])

    label_cell(ws, 33, 1, "Railing LF")
    input_cell(ws, 33, 2, 40, number_format="0")
    ws["C33"] = "linear ft"; ws["C33"].font = label_font()

    label_cell(ws, 34, 1, "Fascia LF")
    input_cell(ws, 34, 2, 70, number_format="0")
    ws["C34"] = "linear ft"; ws["C34"].font = label_font()

    label_cell(ws, 35, 1, "Stair Runs")
    input_cell(ws, 35, 2, 1, number_format="0")
    ws["C35"] = "number of stair runs"; ws["C35"].font = label_font()

    label_cell(ws, 36, 1, "Stair Treads")
    input_cell(ws, 36, 2, 6, number_format="0")
    ws["C36"] = "total treads"; ws["C36"].font = label_font()

    label_cell(ws, 37, 1, "Stair Landings")
    input_cell(ws, 37, 2, 0, number_format="0")
    ws["C37"] = "number of landings"; ws["C37"].font = label_font()

    label_cell(ws, 38, 1, "Wraparound Stair")
    input_cell(ws, 38, 2, "No")
    add_dv(ws, "B38", options=db["yes_no_options"])

    # ---- Stain Inputs (rows 40-43) ---------------------------------------
    section_header_at(ws, 40, 1, 3, "Stain Inputs (only used if project type involves staining)")

    label_cell(ws, 41, 1, "Stain SF")
    input_cell(ws, 41, 2, 0, number_format="0")
    ws["C41"] = "SF of surface to stain"; ws["C41"].font = label_font()

    label_cell(ws, 42, 1, "Stain Type")
    stain_types = sorted({sr["type"] for sr in db["stain_rates_per_sf"]})
    input_cell(ws, 42, 2, stain_types[0])
    add_dv(ws, "B42", options=stain_types)

    label_cell(ws, 43, 1, "Coats")
    input_cell(ws, 43, 2, 1, number_format="0")
    add_dv(ws, "B43", options=["1", "2"])

    # ---- Repair Bucket (rows 45-48) --------------------------------------
    section_header_at(ws, 45, 1, 3, "Repair Bucket (only for Stain+Repairs or Resurface)")

    label_cell(ws, 46, 1, "Board Replacements")
    input_cell(ws, 46, 2, 0, number_format="0")
    ws["C46"] = "number of boards"; ws["C46"].font = label_font()

    label_cell(ws, 47, 1, "Joist Repair LF")
    input_cell(ws, 47, 2, 0, number_format="0")
    ws["C47"] = "linear ft of joists"; ws["C47"].font = label_font()

    label_cell(ws, 48, 1, "Hardware Included?")
    input_cell(ws, 48, 2, "No")
    add_dv(ws, "B48", options=db["yes_no_options"])

    # ---- Adders (rows 50-55) ---------------------------------------------
    section_header_at(ws, 50, 1, 3, "Adders (Optional Upgrades)")

    label_cell(ws, 51, 1, "Skirting / Privacy Wall SF")
    input_cell(ws, 51, 2, 0, number_format="0")
    ws["C51"] = "SF"; ws["C51"].font = label_font()

    label_cell(ws, 52, 1, "Lighting Fixtures")
    input_cell(ws, 52, 2, 0, number_format="0")
    ws["C52"] = "count"; ws["C52"].font = label_font()

    label_cell(ws, 53, 1, "Built-In Benches")
    input_cell(ws, 53, 2, 0, number_format="0")
    ws["C53"] = "count"; ws["C53"].font = label_font()

    label_cell(ws, 54, 1, "Privacy Screen LF")
    input_cell(ws, 54, 2, 0, number_format="0")
    ws["C54"] = "linear ft"; ws["C54"].font = label_font()

    label_cell(ws, 55, 1, "Hot Tub Structural Upgrade")
    input_cell(ws, 55, 2, "No")
    add_dv(ws, "B55", options=db["yes_no_options"])

    # ---- Allowances (rows 57-61) -----------------------------------------
    # Cells B58/B59/B60 are FORMULAS that auto-pull project-type-appropriate
    # defaults from the Allowances by Project Type table. User can override
    # by typing a value (which replaces the formula).
    section_header_at(ws, 57, 1, 3, "Allowances (auto-default per project type; override by typing)")

    label_cell(ws, 58, 1, "Permit / Engineering")
    permit_cell = input_cell(ws, 58, 2, None, number_format='"$"#,##0')
    permit_cell.value = (f"=IFERROR(INDEX({PDB['alwbt_permit']}, "
                         f"MATCH({QI.PROJECT_TYPE}, {PDB['alwbt_name']}, 0)), "
                         f"{PDB['alw_permit_def']})")

    label_cell(ws, 59, 1, "Dumpster / Cleanup")
    dump_cell = input_cell(ws, 59, 2, None, number_format='"$"#,##0')
    dump_cell.value = (f"=IFERROR(INDEX({PDB['alwbt_dump']}, "
                       f"MATCH({QI.PROJECT_TYPE}, {PDB['alwbt_name']}, 0)), "
                       f"{PDB['alw_dump_def']})")

    label_cell(ws, 60, 1, "Mobilization")
    mob_cell = input_cell(ws, 60, 2, None, number_format='"$"#,##0')
    mob_cell.value = (f"=IFERROR(INDEX({PDB['alwbt_mob']}, "
                      f"MATCH({QI.PROJECT_TYPE}, {PDB['alwbt_name']}, 0)), "
                      f"{PDB['alw_mob_def']})")

    label_cell(ws, 61, 1, "Manual Misc Add")
    input_cell(ws, 61, 2, 0, number_format='"$"#,##0')

    # ---- Margin & Contingency (rows 63-65) -------------------------------
    section_header_at(ws, 63, 1, 3, "Margin & Range")
    mc = db["margin_and_contingency"]

    label_cell(ws, 64, 1, "Target Gross Margin")
    input_cell(ws, 64, 2, mc["default_margin"], number_format='0.0%')

    label_cell(ws, 65, 1, "Contingency Range (+/-)")
    input_cell(ws, 65, 2, mc["default_contingency"], number_format='0.0%')

    # ---- Footer notes (rows 67+) -----------------------------------------
    ws["A67"] = "Notes:"
    ws["A67"].font = Font(name="Calibri", size=10, bold=True, color=GREY)
    ws["A68"] = ("- Yellow cells are inputs. Grey cells are formulas. Do not "
                 "type over grey cells.")
    ws["A69"] = ("- Switch Project Type first. Inputs that don't apply to the "
                 "selected type are still editable but the engine ignores them.")
    ws["A70"] = ("- See 'Notes' sheet for the project-type behavior matrix "
                 "and full assumptions.")
    for r in (68, 69, 70):
        ws[f"A{r}"].font = Font(name="Calibri", size=10, color=GREY, italic=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)

    # ---- Fence & Colors (rows 71-81) -------------------------------------
    section_header_at(ws, 71, 1, 3, "Fence Inputs (used only when Project Type = Fence) + Colors")

    label_cell(ws, 72, 1, "Fence Type")
    input_cell(ws, 72, 2, db["fence_materials"][0]["name"])
    add_dv(ws, "B72", options=[m["name"] for m in db["fence_materials"]])

    label_cell(ws, 73, 1, "Fence Height (ft)")
    input_cell(ws, 73, 2, "6")
    add_dv(ws, "B73", options=db.get("fence_heights", ["4", "5", "6", "8"]))

    label_cell(ws, 74, 1, "Fence LF")
    input_cell(ws, 74, 2, 0, number_format="0")

    label_cell(ws, 75, 1, "Walk Gates")
    input_cell(ws, 75, 2, 0, number_format="0")

    label_cell(ws, 76, 1, "Drive / Double Gates")
    input_cell(ws, 76, 2, 0, number_format="0")

    label_cell(ws, 77, 1, "Existing Fence Tear-out LF")
    input_cell(ws, 77, 2, 0, number_format="0")

    label_cell(ws, 78, 1, "Fence Color")
    input_cell(ws, 78, 2, "")

    label_cell(ws, 79, 1, "Decking Color")
    input_cell(ws, 79, 2, "")

    label_cell(ws, 80, 1, "Railing Color")
    input_cell(ws, 80, 2, "")

    label_cell(ws, 81, 1, "Stain Color")
    input_cell(ws, 81, 2, "")

    # =====================================================================
    # Conditional formatting: grey out inputs the engine ignores
    # =====================================================================
    # Stain inputs (rows 41-43) - grey out when matrix.Stain != "Y"
    na_fill = PatternFill("solid", fgColor=NA_GREY)
    stain_rule = FormulaRule(
        formula=[f'INDEX({PB["H"]}, '
                 f'MATCH({QI.PROJECT_TYPE}, {PB["A"]}, 0))<>"Y"'],
        fill=na_fill)
    ws.conditional_formatting.add("A41:C43", stain_rule)
    ws.conditional_formatting.add("D26:E26", stain_rule)

    # Repair inputs (rows 46-48) - grey out when project type isn't stain+repairs or resurface
    repair_rule = FormulaRule(
        formula=[f'AND({QI.PROJECT_TYPE}<>"Stain + Minor Repairs", '
                 f'{QI.PROJECT_TYPE}<>"Resurface (Boards Only)")'],
        fill=na_fill)
    ws.conditional_formatting.add("A46:C48", repair_rule)
    ws.conditional_formatting.add("D27:E27", repair_rule)

    # Scope detail block (rows 32-38) - grey out for Stain Only
    scope_rule = FormulaRule(
        formula=[f'{QI.PROJECT_TYPE}="Stain Only"'],
        fill=na_fill)
    ws.conditional_formatting.add("A32:C38", scope_rule)
    ws.conditional_formatting.add("A51:C55", scope_rule)
    ws.conditional_formatting.add("D22:E25", scope_rule)
    ws.conditional_formatting.add("D28:E32", scope_rule)

    # Materials block - grey out for Stain Only (decking + railing + framing N/A)
    mat_rule = FormulaRule(
        formula=[f'{QI.PROJECT_TYPE}="Stain Only"'],
        fill=na_fill)
    ws.conditional_formatting.add("A21:C23", mat_rule)
    ws.conditional_formatting.add("D7:E10", mat_rule)

    # Freeze panes so the dashboard stays visible while scrolling
    ws.freeze_panes = "A6"

    return ws


# ---------------------------------------------------------------------------
# Build Materials List sheet
# ---------------------------------------------------------------------------
def build_materials_list(wb, db):
    """
    Preliminary materials shopping list. Quantities derived from Quote Input
    dimensions + project type matrix; unit costs from Pricing DB Materials
    Unit Costs table (rows 130-157).
    """
    ws = wb.create_sheet("Materials List")

    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 38

    # ---------- helpers --------------------------------------------------
    def mat_matrix(col_letter):
        """Cross-sheet matrix lookup for current project type."""
        return (f"INDEX({PB[col_letter]}, "
                f"MATCH('Quote Input'!{QI.PROJECT_TYPE}, "
                f"{PB['A']}, 0))")

    def uc(key):
        """Unit cost lookup against Pricing DB Materials Unit Costs."""
        return (f"INDEX({PDB['muc_price']}, "
                f"MATCH(\"{key}\", {PDB['muc_key']}, 0))")

    def deck_board_unit_cost():
        """Per-board cost = material $/SF × 7.33 SF/board (5/4x6x16)."""
        return (f"INDEX({PDB['dk_mat']}, "
                f"MATCH('Quote Input'!{QI.DECKING_MAT}, "
                f"{PDB['dk_name']}, 0))*7.33")

    def section(row, text):
        section_header_at(ws, row, 1, 6, text)

    def table_header_row(row):
        for i, h in enumerate(["Item", "Qty", "Unit", "Unit Cost",
                                "Line Cost", "Notes"], start=1):
            header_cell(ws, row, i, h)

    def item_row(row, label, qty_formula, unit, unit_cost_formula, notes):
        a = ws.cell(row=row, column=1)
        a.value = label
        a.font = label_font()
        a.alignment = LEFT

        b = ws.cell(row=row, column=2)
        b.value = qty_formula
        b.number_format = "0"
        b.alignment = RIGHT
        b.font = label_font()

        c = ws.cell(row=row, column=3)
        c.value = unit
        c.font = label_font()
        c.alignment = CENTER

        d = ws.cell(row=row, column=4)
        d.value = unit_cost_formula
        d.number_format = '"$"#,##0.00'
        d.alignment = RIGHT
        d.font = calc_font()
        d.fill = fill(CALC_GREY)

        e = ws.cell(row=row, column=5)
        e.value = f"=B{row}*D{row}"
        e.number_format = '"$"#,##0'
        e.alignment = RIGHT
        e.font = label_font(bold=True)

        f = ws.cell(row=row, column=6)
        f.value = notes
        f.font = Font(name="Calibri", size=9, italic=True, color=GREY)
        f.alignment = LEFT

    # ---------- title and project echo -----------------------------------
    ws["A1"] = "Preliminary Materials List"
    ws["A1"].font = title_font()
    ws.merge_cells("A1:F1")

    ws["A2"] = ("Shopping list driven by Quote Input. Quantities update live. "
                "Edit unit costs in Pricing DB > Materials Unit Costs section. "
                "Total at bottom is materials only - labor and margin live in the Quote.")
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=GREY)
    ws["A2"].alignment = WRAP_LEFT
    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = 32

    ws["A4"] = "Project Type:"
    ws["A4"].font = label_font(bold=True)
    ws["B4"] = f"='Quote Input'!{QI.PROJECT_TYPE}"
    ws["B4"].font = label_font()
    ws.merge_cells("B4:F4")

    ws["A5"] = "Dimensions:"
    ws["A5"].font = label_font(bold=True)
    ws["B5"] = (f"='Quote Input'!{QI.LENGTH}&\" ft x \"&"
                f"'Quote Input'!{QI.DEPTH}&\" ft = \"&"
                f"'Quote Input'!{QI.DECK_SF}&\" SF deck\"")
    ws["B5"].font = label_font()
    ws.merge_cells("B5:F5")

    ws["A6"] = "Materials selected:"
    ws["A6"].font = label_font(bold=True)
    ws["B6"] = (f"=\"Decking: \"&'Quote Input'!{QI.DECKING_MAT}&"
                f"\"  |  Railing: \"&'Quote Input'!{QI.RAILING_MAT}&"
                f"\"  |  Framing: \"&'Quote Input'!{QI.FRAMING_MAT}")
    ws["B6"].font = label_font()
    ws.merge_cells("B6:F6")

    # ---------- DECKING --------------------------------------------------
    section(8, "Decking")
    table_header_row(9)

    # Row 10: Deck boards (matrix.deck=Y)
    item_row(10,
        label=f"=\"Deck board: \"&'Quote Input'!{QI.DECKING_MAT}&\" (5/4x6x16)\"",
        qty_formula=(f"=IF({mat_matrix('D')}=\"Y\", "
                     f"CEILING('Quote Input'!{QI.DECK_SF}*"
                     f"(1+'Quote Input'!{QI.DECKING_WASTE})/7.33, 1), 0)"),
        unit="each",
        unit_cost_formula=f"=IFERROR({deck_board_unit_cost()}, 0)",
        notes="5/4x6x16 nominal = 5.5\" face x 16 ft = 7.33 SF/board"
    )

    # Row 11: Hidden fasteners (composite only)
    item_row(11,
        label=f"=IF(LEFT('Quote Input'!{QI.DECKING_MAT},4)=\"Trex\", "
              f"\"Hidden fasteners (composite) - 50 ct box\", "
              f"\"3-inch stainless deck screws - 5 lb box\")",
        qty_formula=(f"=IF({mat_matrix('D')}<>\"Y\", 0, "
                     f"IF(LEFT('Quote Input'!{QI.DECKING_MAT},4)=\"Trex\", "
                     f"CEILING('Quote Input'!{QI.DECK_SF}/25, 1), "
                     f"CEILING('Quote Input'!{QI.DECK_SF}/100, 1)))"),
        unit="box",
        unit_cost_formula=(f"=IF(LEFT('Quote Input'!{QI.DECKING_MAT},4)=\"Trex\", "
                           f"{uc('hidden_fastener_box')}, "
                           f"{uc('deck_screw_box_5lb')})"),
        notes="Composite: ~1 box / 25 SF. Wood deck: ~1 box (5 lb) per 100 SF"
    )

    # Row 12: Fascia boards
    item_row(12,
        label=f"=\"Fascia board: \"&'Quote Input'!{QI.DECKING_MAT}&\" (16 ft)\"",
        qty_formula=(f"=IF({mat_matrix('D')}<>\"Y\", 0, "
                     f"CEILING('Quote Input'!{QI.FASCIA_LF}/16, 1))"),
        unit="each",
        unit_cost_formula=f"=IFERROR({deck_board_unit_cost()}, 0)",
        notes="Use matching deck material on exposed rim and stair stringers"
    )

    # Row 13: Construction adhesive
    item_row(13,
        label="Construction adhesive (PL Premium 10 oz)",
        qty_formula=(f"=IF({mat_matrix('D')}<>\"Y\", 0, "
                     f"CEILING('Quote Input'!{QI.DECK_SF}/100, 1))"),
        unit="tube",
        unit_cost_formula=f"={uc('construction_adhesive')}",
        notes="~1 tube per 100 SF (ledger, stair stringers, beam laminations)"
    )

    # ---------- FRAMING --------------------------------------------------
    section(15, "Framing - KDAT Pressure-Treated (rows 17-23 zero out if Steel framing selected; see row 57)")
    table_header_row(16)

    # Wood framing rows fire only when matrix.frame=Y AND framing_mat=KDAT
    wood_frame_active = (f"AND({mat_matrix('C')}=\"Y\", "
                         f"'Quote Input'!{QI.FRAMING_MAT}=\"KDAT Pressure-Treated\")")

    # Row 17: Joists 2x10x16 (count = length x 12 / 16 spacing + 1)
    item_row(17,
        label="KDAT PT 2x10x16 (joists, 16\" o.c. spacing)",
        qty_formula=(f"=IF(NOT({wood_frame_active}), 0, "
                     f"CEILING('Quote Input'!{QI.LENGTH}*12/16, 1)+1)"),
        unit="each",
        unit_cost_formula=f"={uc('kdat_2x10x16')}",
        notes="One joist every 16 in along the deck length, plus 1 end joist"
    )

    # Row 18: Beam (2-ply 2x10x16) - 2 boards per beam-length section
    item_row(18,
        label="KDAT PT 2x10x16 (beam, 2-ply)",
        qty_formula=(f"=IF(NOT({wood_frame_active}), 0, "
                     f"2*CEILING('Quote Input'!{QI.LENGTH}/16, 1))"),
        unit="each",
        unit_cost_formula=f"={uc('kdat_2x10x16')}",
        notes="2-ply doubled beam under outer joist ends; lengths cut to span"
    )

    # Row 19: Ledger
    item_row(19,
        label="KDAT PT 2x10x16 (ledger to house)",
        qty_formula=(f"=IF(NOT({wood_frame_active}), 0, "
                     f"CEILING('Quote Input'!{QI.LENGTH}/16, 1))"),
        unit="each",
        unit_cost_formula=f"={uc('kdat_2x10x16')}",
        notes="Lag-bolted to house rim joist; sealed with flashing"
    )

    # Row 20: Posts 6x6x8 (one per ~6 ft of beam, plus 1)
    item_row(20,
        label="KDAT PT 6x6x8 (beam posts)",
        qty_formula=(f"=IF(NOT({wood_frame_active}), 0, "
                     f"CEILING('Quote Input'!{QI.LENGTH}/6, 1)+1)"),
        unit="each",
        unit_cost_formula=f"={uc('kdat_6x6x8')}",
        notes="Spaced ~6 ft on center along beam; 8-ft length cut to height"
    )

    # Row 21: Joist hangers (count = joist count = joist row qty)
    item_row(21,
        label="Joist hangers (Simpson LUS28 for 2x10)",
        qty_formula=(f"=IF(NOT({wood_frame_active}), 0, "
                     f"CEILING('Quote Input'!{QI.LENGTH}*12/16, 1)+1)"),
        unit="each",
        unit_cost_formula=f"={uc('joist_hanger_2x10')}",
        notes="One per joist at ledger-end and beam-end (count matches joists)"
    )

    # Row 22: Hurricane ties
    item_row(22,
        label="Hurricane ties (Simpson H1)",
        qty_formula=(f"=IF(NOT({wood_frame_active}), 0, "
                     f"CEILING('Quote Input'!{QI.LENGTH}*12/16, 1)+1)"),
        unit="each",
        unit_cost_formula=f"={uc('hurricane_tie')}",
        notes="One per joist where it crosses beam (code-required)"
    )

    # Row 23: Ledger lag screws
    item_row(23,
        label="1/2\" x 4\" lag screws (ledger to rim)",
        qty_formula=(f"=IF(NOT({wood_frame_active}), 0, "
                     f"CEILING('Quote Input'!{QI.LENGTH}*12/16, 1))"),
        unit="each",
        unit_cost_formula=f"={uc('ledger_lag_screw')}",
        notes="One every 16\" along ledger (Wisconsin code)"
    )

    # ---------- FOOTINGS -------------------------------------------------
    section(25, "Footings (matrix.footing = Y only)")
    table_header_row(26)

    # Posts count formula reused
    post_count = (f"(CEILING('Quote Input'!{QI.LENGTH}/6, 1)+1)")

    # Row 27: Tube forms
    item_row(27,
        label="Concrete tube form 12\" x 48\" (WI frost line)",
        qty_formula=f"=IF({mat_matrix('G')}<>\"Y\", 0, {post_count})",
        unit="each",
        unit_cost_formula=f"={uc('concrete_tube_12x48')}",
        notes="One tube per post; 48\" depth for Wisconsin frost line"
    )

    # Row 28: Concrete bags
    item_row(28,
        label="Concrete mix 60 lb bag",
        qty_formula=f"=IF({mat_matrix('G')}<>\"Y\", 0, {post_count}*4)",
        unit="bag",
        unit_cost_formula=f"={uc('concrete_bag_60lb')}",
        notes="~4 bags per 12\" x 48\" tube footing"
    )

    # Row 29: Rebar
    item_row(29,
        label="Rebar 1/2\" x 4 ft",
        qty_formula=f"=IF({mat_matrix('G')}<>\"Y\", 0, {post_count}*2)",
        unit="each",
        unit_cost_formula=f"={uc('rebar_half_x_4ft')}",
        notes="2 pieces per footing in cross pattern"
    )

    # ---------- RAILING --------------------------------------------------
    section(31, "Railing")
    table_header_row(32)

    rail_active = (f"AND({mat_matrix('E')}<>\"N\", "
                   f"'Quote Input'!{QI.RAILING_LF}>0)")

    is_trex_alum = f"'Quote Input'!{QI.RAILING_MAT}=\"Trex Aluminum\""
    is_pt_wood = f"'Quote Input'!{QI.RAILING_MAT}=\"Pressure-Treated Wood Rail\""
    is_cedar_wood = f"'Quote Input'!{QI.RAILING_MAT}=\"Cedar Wood Rail\""
    is_wood = f"OR({is_pt_wood}, {is_cedar_wood})"
    is_cable = f"'Quote Input'!{QI.RAILING_MAT}=\"Cable Railing\""
    is_glass = f"'Quote Input'!{QI.RAILING_MAT}=\"Glass Panel\""

    # Row 33: Aluminum rail section (Trex Aluminum only)
    item_row(33,
        label="Trex Signature aluminum rail section (6 ft kit)",
        qty_formula=(f"=IF(AND({rail_active}, {is_trex_alum}), "
                     f"CEILING('Quote Input'!{QI.RAILING_LF}/6, 1), 0)"),
        unit="kit",
        unit_cost_formula=f"={uc('trex_alum_rail_6ft')}",
        notes="Includes top/bot rail, balusters; cut to length on site"
    )

    # Row 34: Aluminum post kits
    item_row(34,
        label="Trex Signature aluminum post + base kit",
        qty_formula=(f"=IF(AND({rail_active}, {is_trex_alum}), "
                     f"CEILING('Quote Input'!{QI.RAILING_LF}/6, 1)+1, 0)"),
        unit="kit",
        unit_cost_formula=f"={uc('trex_alum_post_kit')}",
        notes="One post per 6 ft of railing, plus 1 for the end"
    )

    # Row 35: Wood top + bottom rail (PT or Cedar)
    item_row(35,
        label=f"=IF({is_pt_wood}, \"AC2 PT 2x4x8 \", IF({is_cedar_wood}, "
              f"\"Cedar 2x4x8 \", \"\"))&\"(top + bottom rail)\"",
        qty_formula=(f"=IF(AND({rail_active}, {is_wood}), "
                     f"2*CEILING('Quote Input'!{QI.RAILING_LF}/8, 1), 0)"),
        unit="each",
        unit_cost_formula=(f"=IF({is_pt_wood}, {uc('pt_2x4x8')}, "
                           f"IF({is_cedar_wood}, {uc('cedar_2x4x8')}, 0))"),
        notes="2 boards (top + bottom) per 8 LF of railing"
    )

    # Row 36: Wood balusters
    item_row(36,
        label=f"=IF({is_pt_wood}, \"PT 2x2x42 baluster\", IF({is_cedar_wood}, "
              f"\"Cedar 2x2x42 baluster\", \"\"))",
        qty_formula=(f"=IF(AND({rail_active}, {is_wood}), "
                     f"CEILING('Quote Input'!{QI.RAILING_LF}*12/4, 1), 0)"),
        unit="each",
        unit_cost_formula=(f"=IF({is_pt_wood}, {uc('pt_2x2x42')}, "
                           f"IF({is_cedar_wood}, {uc('cedar_2x2x42')}, 0))"),
        notes="One baluster every 4 in (code-required spacing)"
    )

    # Row 37: Wood rail posts (4x4 PT - same for both PT/cedar systems)
    item_row(37,
        label="KDAT PT 4x4x8 (rail posts)",
        qty_formula=(f"=IF(AND({rail_active}, {is_wood}), "
                     f"CEILING('Quote Input'!{QI.RAILING_LF}/6, 1)+1, 0)"),
        unit="each",
        unit_cost_formula=f"={uc('kdat_4x4x8')}",
        notes="Wood rail post every 6 ft; PT used even on cedar rail systems"
    )

    # Row 38: Cable kits
    item_row(38,
        label="Feeney CableRail 10 ft kit",
        qty_formula=(f"=IF(AND({rail_active}, {is_cable}), "
                     f"CEILING('Quote Input'!{QI.RAILING_LF}/10, 1), 0)"),
        unit="kit",
        unit_cost_formula=f"={uc('cable_rail_kit_10ft')}",
        notes="10 ft cable assembly + tensioners; need wood posts (see row 37)"
    )

    # Row 39: Glass panels
    item_row(39,
        label="Tempered glass railing panel 36\" x 42\"",
        qty_formula=(f"=IF(AND({rail_active}, {is_glass}), "
                     f"CEILING('Quote Input'!{QI.RAILING_LF}/3, 1), 0)"),
        unit="panel",
        unit_cost_formula=f"={uc('glass_panel_36x42')}",
        notes="3 ft panels with aluminum posts every 3 ft"
    )

    # Row 40: Glass posts
    item_row(40,
        label="Aluminum glass-railing post + clamp set",
        qty_formula=(f"=IF(AND({rail_active}, {is_glass}), "
                     f"CEILING('Quote Input'!{QI.RAILING_LF}/3, 1)+1, 0)"),
        unit="kit",
        unit_cost_formula=f"={uc('alum_glass_post')}",
        notes="One per panel edge + 1 for the end"
    )

    # ---------- STAIRS ---------------------------------------------------
    section(42, "Stairs (only if stair runs > 0)")
    table_header_row(43)

    stair_active = (f"AND({mat_matrix('F')}<>\"N\", "
                    f"'Quote Input'!{QI.STAIR_RUNS}>0)")

    # Row 44: Stringers (2 per stair run; use 2x10x16 to leave cutting room)
    item_row(44,
        label="KDAT PT 2x10x16 (stair stringers, cut from blanks)",
        qty_formula=f"=IF({stair_active}, 'Quote Input'!{QI.STAIR_RUNS}*2, 0)",
        unit="each",
        unit_cost_formula=f"={uc('kdat_2x10x16')}",
        notes="2 stringers per stair run; cut from 16 ft blank"
    )

    # Row 45: Tread boards (count using deck material)
    item_row(45,
        label=f"=\"Tread board: \"&'Quote Input'!{QI.DECKING_MAT}",
        qty_formula=(f"=IF({stair_active}, "
                     f"CEILING('Quote Input'!{QI.STAIR_TREADS}/2, 1), 0)"),
        unit="each",
        unit_cost_formula=f"=IFERROR({deck_board_unit_cost()}, 0)",
        notes="2 tread pieces per board (16 ft cut into 4 x 4 ft pieces, used as 2 treads)"
    )

    # Row 46: Riser boards
    item_row(46,
        label="KDAT PT 1x8 riser stock (or matching deck material)",
        qty_formula=(f"=IF({stair_active}, "
                     f"CEILING('Quote Input'!{QI.STAIR_TREADS}/2, 1), 0)"),
        unit="each",
        unit_cost_formula=f"={uc('kdat_2x10x16')}*0.5",
        notes="Vertical face of each stair; cut from 16 ft stock"
    )

    # Row 47: Stair hardware (lump)
    item_row(47,
        label="Stair hardware (brackets, screws, drip caps)",
        qty_formula=f"=IF({stair_active}, 'Quote Input'!{QI.STAIR_RUNS}, 0)",
        unit="set",
        unit_cost_formula="=85",
        notes="~$85 per stair run for misc fasteners + brackets"
    )

    # ---------- STAIN ----------------------------------------------------
    section(49, "Stain (only if matrix.stain = Y)")
    table_header_row(50)

    stain_active = f"AND({mat_matrix('H')}=\"Y\", 'Quote Input'!{QI.STAIN_SF}>0)"

    # Row 51: Stain gallons
    item_row(51,
        label=f"=\"Stain (\"&'Quote Input'!{QI.STAIN_TYPE}&\") - 1 gal\"",
        qty_formula=(f"=IF({stain_active}, "
                     f"CEILING('Quote Input'!{QI.STAIN_SF}*"
                     f"'Quote Input'!{QI.STAIN_COATS}/300, 1), 0)"),
        unit="gal",
        unit_cost_formula=(f"=IF('Quote Input'!{QI.STAIN_TYPE}=\"Transparent\", "
                           f"{uc('stain_transparent_gal')}, "
                           f"IF('Quote Input'!{QI.STAIN_TYPE}=\"Semi-Transparent\", "
                           f"{uc('stain_semi_gal')}, {uc('stain_solid_gal')}))"),
        notes="~300 SF/gallon coverage; multiply by coat count"
    )

    # Row 52: Stain supplies set
    item_row(52,
        label="Stain supplies set (brushes, rollers, pads, drop cloths)",
        qty_formula=f"=IF({stain_active}, 1, 0)",
        unit="set",
        unit_cost_formula=f"={uc('stain_supplies_set')}",
        notes="Consumable supplies for the job; one set per project"
    )

    # ---------- OTHER ----------------------------------------------------
    section(54, "Other / Site")
    table_header_row(55)

    # Row 56: Dumpster
    item_row(56,
        label="10-yard roll-off dumpster (1 week)",
        qty_formula=(f"=IF({mat_matrix('B')}=\"N\", 0, 1)"),
        unit="rental",
        unit_cost_formula=f"={uc('dumpster_rental')}",
        notes="On-site dumpster for demo debris and offcuts; bigger for full tear-out"
    )

    # Row 57: Steel framing system (Fortress Evolution) - alternative to lumber
    item_row(57,
        label="Fortress Evolution steel framing system",
        qty_formula=(f"=IF(AND({mat_matrix('C')}=\"Y\", "
                     f"'Quote Input'!{QI.FRAMING_MAT}=\"Steel (Fortress Evolution)\"), "
                     f"'Quote Input'!{QI.DECK_SF}, 0)"),
        unit="SF",
        unit_cost_formula=f"={uc('steel_framing_per_sf')}",
        notes="REPLACES rows 17-20 lumber when steel framing is selected; "
              "those wood rows will show 0 in that case"
    )

    # ---------- TOTAL ----------------------------------------------------
    section(59, "TOTAL MATERIALS COST")
    ws["A60"] = "Estimated materials cost (Menards Wausau + supplier prices):"
    ws["A60"].font = Font(name="Calibri", size=11, bold=True, color=SLATE)
    ws.merge_cells("A60:D60")
    ws["E60"] = "=SUM(E10:E57)"
    ws["E60"].font = Font(name="Calibri", size=14, bold=True, color=ORANGE)
    ws["E60"].number_format = '"$"#,##0'
    ws["E60"].fill = fill(LIGHT_BG)
    ws["E60"].alignment = RIGHT

    ws["A61"] = ("Note: this is the materials shopping list only. Labor, "
                 "overhead, and gross margin are computed separately in the "
                 "Quote Input tab. Materials cost / Quote Sell Price gives "
                 "a rough materials ratio.")
    ws["A61"].font = Font(name="Calibri", size=9, italic=True, color=GREY)
    ws["A61"].alignment = WRAP_LEFT
    ws.merge_cells("A61:F61")
    ws.row_dimensions[61].height = 32

    # ---------- conditional formatting -----------------------------------
    # Grey out rows where Qty = 0 (item not applicable to current project type)
    na_fill = PatternFill("solid", fgColor=NA_GREY)
    # Apply to each item row in turn
    item_rows = [10, 11, 12, 13, 17, 18, 19, 20, 21, 22, 23,
                 27, 28, 29, 33, 34, 35, 36, 37, 38, 39, 40,
                 44, 45, 46, 47, 51, 52, 56, 57]
    for r in item_rows:
        rule = FormulaRule(formula=[f"$B{r}=0"], fill=na_fill)
        ws.conditional_formatting.add(f"A{r}:F{r}", rule)

    # Freeze the project info echo (rows 1-7) so it stays visible while scrolling
    ws.freeze_panes = "A8"

    return ws


# ---------------------------------------------------------------------------
# Build Estimate Summary sheet
# ---------------------------------------------------------------------------
def build_estimate_summary(wb, db):
    """Three-tier decking comparison (Trex Enhance / Select / Transcend) when
    project type includes a new deck. For stain-only or stain+repairs, shows
    a single-quote summary instead."""
    ws = wb.create_sheet("Estimate Summary")

    ws.column_dimensions["A"].width = 26
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 18

    ws["A1"] = "Estimate Summary"
    ws["A1"].font = title_font()
    ws.merge_cells("A1:H1")

    ws["A2"] = ("Three-tier comparison for new-deck projects. Single-tier "
                "summary for stain/repair jobs.")
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=GREY)
    ws.merge_cells("A2:H2")

    # Project info echo
    section_header(ws, 4, "Project", span=8)
    label_cell(ws, 5, 1, "Client")
    ws["B5"] = f"='Quote Input'!{QI.CLIENT_NAME}"
    label_cell(ws, 6, 1, "Address")
    ws["B6"] = f"='Quote Input'!{QI.ADDRESS}"
    label_cell(ws, 7, 1, "Project Type")
    ws["B7"] = f"='Quote Input'!{QI.PROJECT_TYPE}"
    label_cell(ws, 8, 1, "Deck SF")
    ws["B8"] = f"='Quote Input'!{QI.DECK_SF}"
    label_cell(ws, 9, 1, "Selected Decking")
    ws["B9"] = f"='Quote Input'!{QI.DECKING_MAT}"

    # Three-tier comparison table
    section_header(ws, 11, "Three-Tier Decking Comparison (composite only)", span=8)
    headers = ["Decking", "Decking $/SF", "Waste", "Base Package",
               "Adders (shared)", "Subtotal", "Sell Price", "Range"]
    for i, h in enumerate(headers, start=1):
        header_cell(ws, 12, i, h)

    # Adders shared across tiers (everything except base package)
    adders_formula = (f"('Quote Input'!{QI.DEMO_COST}+"
                      f"'Quote Input'!{QI.BORDER_COST}+"
                      f"'Quote Input'!{QI.RAILING_COST}+"
                      f"'Quote Input'!{QI.FASCIA_COST}+"
                      f"'Quote Input'!{QI.STAIR_COST}+"
                      f"'Quote Input'!{QI.STAIN_COST}+"
                      f"'Quote Input'!{QI.REPAIR_COST}+"
                      f"'Quote Input'!{QI.SKIRTING_COST}+"
                      f"'Quote Input'!{QI.LIGHTING_COST}+"
                      f"'Quote Input'!{QI.BENCH_COST}+"
                      f"'Quote Input'!{QI.PRIV_SCREEN_COST}+"
                      f"'Quote Input'!{QI.HOT_TUB_COST}+"
                      f"'Quote Input'!{QI.PERMIT_COST}+"
                      f"'Quote Input'!{QI.DUMPSTER_COST}+"
                      f"'Quote Input'!{QI.MOBIL_COST}+"
                      f"'Quote Input'!{QI.MISC_COST})")

    composite_names = ["Trex Enhance", "Trex Select", "Trex Transcend"]
    for i, name in enumerate(composite_names):
        r = 13 + i
        db_cell(ws, r, 1, name, editable=False)

        # Decking $/SF
        decking_psf = (f"IFERROR(INDEX({PDB['dk_sell']}, "
                       f"MATCH(\"{name}\", {PDB['dk_name']}, 0)), 0)")
        ws.cell(row=r, column=2).value = f"={decking_psf}"
        ws.cell(row=r, column=2).number_format = '"$"#,##0.00'

        # Waste %
        waste = (f"IFERROR(INDEX({PDB['dk_waste']}, "
                 f"MATCH(\"{name}\", {PDB['dk_name']}, 0)), 0)")
        ws.cell(row=r, column=3).value = f"={waste}"
        ws.cell(row=r, column=3).number_format = '0.0%'

        # Base Package = DeckSF * (Framing + Decking*(1+waste)) * Multiplier
        # (matrix-aware: only if frame/deck included)
        base = (f"='Quote Input'!{QI.DECK_SF}*"
                f"(IF(INDEX({PB['C']}, "
                f"MATCH('Quote Input'!{QI.PROJECT_TYPE}, "
                f"{PB['A']}, 0))=\"Y\", "
                f"'Quote Input'!{QI.FRAMING_PSF}, 0) + "
                f"IF(INDEX({PB['D']}, "
                f"MATCH('Quote Input'!{QI.PROJECT_TYPE}, "
                f"{PB['A']}, 0))=\"Y\", "
                f"{decking_psf}*(1+{waste}), 0))"
                f"*'Quote Input'!{QI.COMBINED_MULT}")
        ws.cell(row=r, column=4).value = base
        ws.cell(row=r, column=4).number_format = '"$"#,##0'

        # Adders
        ws.cell(row=r, column=5).value = f"={adders_formula}"
        ws.cell(row=r, column=5).number_format = '"$"#,##0'

        # Subtotal
        ws.cell(row=r, column=6).value = (f"=D{r}+E{r}")
        ws.cell(row=r, column=6).number_format = '"$"#,##0'

        # Sell Price
        ws.cell(row=r, column=7).value = (f"=F{r}/(1-'Quote Input'!{QI.TARGET_MARGIN})")
        ws.cell(row=r, column=7).number_format = '"$"#,##0'
        ws.cell(row=r, column=7).font = Font(name="Calibri", size=10, bold=True, color=ORANGE)

        # Range
        ws.cell(row=r, column=8).value = (f'=TEXT(G{r}*(1-\'Quote Input\'!{QI.CONTINGENCY}),'
                                          f'"$#,##0")&" - "&'
                                          f'TEXT(G{r}*(1+\'Quote Input\'!{QI.CONTINGENCY}),'
                                          f'"$#,##0")')

    # Selected quote echo
    section_header(ws, 17, "Selected Quote (from Quote Input)", span=8)
    label_cell(ws, 18, 1, "Sell Price", bold=True)
    ws["B18"] = f"='Quote Input'!{QI.SELL_PRICE}"
    ws["B18"].number_format = '"$"#,##0'
    ws["B18"].font = Font(name="Calibri", size=14, bold=True, color=ORANGE)
    label_cell(ws, 19, 1, "Range")
    ws["B19"] = (f'=TEXT(\'Quote Input\'!{QI.LOW_RANGE},"$#,##0")&" - "&'
                 f'TEXT(\'Quote Input\'!{QI.HIGH_RANGE},"$#,##0")')
    label_cell(ws, 20, 1, "Subtotal Cost")
    ws["B20"] = f"='Quote Input'!{QI.SUBTOTAL}"
    ws["B20"].number_format = '"$"#,##0'
    label_cell(ws, 21, 1, "Margin")
    ws["B21"] = f"='Quote Input'!{QI.TARGET_MARGIN}"
    ws["B21"].number_format = '0.0%'

    return ws


# ---------------------------------------------------------------------------
# Build Client Range sheet
# ---------------------------------------------------------------------------
def build_client_range(wb, db):
    ws = wb.create_sheet("Client Range")
    ws.column_dimensions["A"].width = 110

    ws["A1"] = "Client-Facing Preliminary Range"
    ws["A1"].font = title_font()

    ws["A2"] = ("Copy/paste the block below into a proposal email. "
                "All values reference Quote Input live.")
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=GREY)

    section_header_at(ws, 4, 1, 1, "Proposal Block")

    template = (
        f'="Dear "&\'Quote Input\'!{QI.CLIENT_NAME}&","&CHAR(10)&CHAR(10)&'
        f'"Thank you for the opportunity to quote your deck project. "&'
        f'"Based on the project scope at "&\'Quote Input\'!{QI.ADDRESS}&", "&'
        f'"we estimate your preliminary investment range at:"&CHAR(10)&CHAR(10)&'
        f'TEXT(\'Quote Input\'!{QI.LOW_RANGE},"$#,##0")&" - "&'
        f'TEXT(\'Quote Input\'!{QI.HIGH_RANGE},"$#,##0")&CHAR(10)&CHAR(10)&'
        f'"Project type: "&\'Quote Input\'!{QI.PROJECT_TYPE}&CHAR(10)&'
        f'"Approximate size: "&\'Quote Input\'!{QI.DECK_SF}&" sq ft"&CHAR(10)&'
        f'"Decking: "&\'Quote Input\'!{QI.DECKING_MAT}&CHAR(10)&'
        f'"Railing: "&\'Quote Input\'!{QI.RAILING_MAT}&CHAR(10)&CHAR(10)&'
        f'"This is a preliminary range based on stated scope. A firm fixed-"&'
        f'"price quote will follow site inspection and final scope review. "&'
        f'"Pricing assumes standard site access and conditions; engineering, "&'
        f'"permits, and unusual site conditions may adjust final pricing."'
        f'&CHAR(10)&CHAR(10)&'
        f'"James Slogar, Owner"&CHAR(10)&'
        f'"Central Wisconsin Deck Builders, LLC"&CHAR(10)&'
        f'"(715) 544-7941  ·  cwdeckbuilders.com"'
    )
    ws["A5"] = template
    ws["A5"].font = Font(name="Calibri", size=10, color=SLATE)
    ws["A5"].alignment = Alignment(horizontal="left", vertical="top",
                                    wrap_text=True)
    ws.row_dimensions[5].height = 400

    return ws


# ---------------------------------------------------------------------------
# Build Scope Builder sheet
# ---------------------------------------------------------------------------
def build_scope_builder(wb, db):
    ws = wb.create_sheet("Scope Builder")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 100

    ws["A1"] = "Scope Boilerplate by Project Type"
    ws["A1"].font = title_font()
    ws.merge_cells("A1:B1")

    ws["A2"] = ("Standard scope, included, and excluded items by project "
                "type. Copy into your client proposal as appropriate.")
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=GREY)
    ws.merge_cells("A2:B2")

    scope_blocks = [
        ("Stain Only", [
            ("Scope of Work",
             "Wash, sand rough spots, set raised fasteners, apply selected "
             "stain product to deck floor, railings, and stairs as specified. "
             "Move and protect homeowner furniture; cover landscaping and "
             "siding. Final touch-ups and full job-site cleanup."),
            ("Included",
             "All materials (stain, brushes, rollers, pads, sandpaper, drop "
             "cloths, masking). All labor. Site protection. 1-year limited "
             "workmanship warranty against premature finish failure."),
            ("Not Included",
             "Repair or replacement of damaged or rotted boards. Hardware "
             "replacement beyond setting existing fasteners. Structural "
             "repairs. Refinishing of pergolas, gazebos, or attached fences "
             "unless added in writing."),
        ]),
        ("Stain + Minor Repairs", [
            ("Scope of Work",
             "Stain Only scope plus targeted repairs: replace damaged deck "
             "boards (count specified), sister or repair compromised joists "
             "(LF specified), replace failed hardware as needed. Repairs "
             "completed before staining."),
            ("Included",
             "Stain Only includes plus: repair-grade replacement boards, "
             "joist sister lumber, fasteners, hidden hardware as needed. "
             "Labor for repairs included. Same 1-year workmanship warranty."),
            ("Not Included",
             "Wholesale frame replacement (use Frame + Deck Rebuild). New "
             "footings. Code upgrades beyond the repair scope."),
        ]),
        ("Resurface (Boards Only)", [
            ("Scope of Work",
             "Remove existing deck boards. Inspect frame for soundness "
             "(replace failed joists/beams as flagged). Install new deck "
             "boards in selected material with hidden fasteners. Refresh "
             "railings and stairs as elected."),
            ("Included",
             "All new decking material, hidden fasteners, fascia (if "
             "specified). Frame inspection allowance. Labor for tear-off "
             "and install. Cleanup and disposal of old boards. 5-year "
             "workmanship warranty on new boards."),
            ("Not Included",
             "Major frame replacement (quoted separately if uncovered "
             "during inspection). New footings. Code-driven structural "
             "upgrades."),
        ]),
        ("Frame + Deck Rebuild (Keep Footings)", [
            ("Scope of Work",
             "Remove existing deck and frame down to footings. Inspect "
             "footings for soundness. Build new code-compliant frame with "
             "selected framing material. Install new deck boards, railings, "
             "stairs, and fascia in selected materials."),
            ("Included",
             "All new framing material (joists, beams, ledger, hangers, "
             "hardware). All new decking, railings, stairs. Permit and "
             "engineering allowance. Demo and disposal. 10-year workmanship "
             "warranty on structure."),
            ("Not Included",
             "New footings (quoted separately if uncovered as inadequate "
             "during inspection). Landscape restoration beyond the deck "
             "footprint. Electrical, plumbing, or hot tub utility hookups."),
        ]),
        ("Full Tear-Out + New Build", [
            ("Scope of Work",
             "Complete demolition of existing deck, frame, and footings. "
             "Layout and install new code-compliant footings. Build new "
             "frame, deck, railings, stairs, and fascia in selected "
             "materials. Final inspection and homeowner walkthrough."),
            ("Included",
             "Permit and engineering allowance. All footings, framing, "
             "decking, railings, stairs. Demo and full disposal. "
             "Site protection. Final walkthrough. 10-year workmanship "
             "warranty on structure. Manufacturer warranties pass through "
             "on decking and railing."),
            ("Not Included",
             "Landscape restoration beyond the deck footprint. Electrical, "
             "plumbing, or hot tub utility hookups. Pergolas, screens, or "
             "outdoor kitchens unless added in writing."),
        ]),
    ]

    r = 4
    for proj_type_name, blocks in scope_blocks:
        # Section header
        cell = ws.cell(row=r, column=1)
        cell.value = proj_type_name
        cell.font = section_font()
        cell.fill = fill(ORANGE)
        ws.cell(row=r, column=2).fill = fill(ORANGE)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        ws.row_dimensions[r].height = 22
        r += 1

        for label, content in blocks:
            ws.cell(row=r, column=1).value = label
            ws.cell(row=r, column=1).font = label_font(bold=True)
            ws.cell(row=r, column=1).alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True)
            ws.cell(row=r, column=2).value = content
            ws.cell(row=r, column=2).font = label_font()
            ws.cell(row=r, column=2).alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True)
            ws.row_dimensions[r].height = max(40, len(content) // 2)
            r += 1
        r += 1  # gap between project types

    return ws


# ---------------------------------------------------------------------------
# Build Notes sheet
# ---------------------------------------------------------------------------
def build_notes(wb, db):
    ws = wb.create_sheet("Notes")
    ws.column_dimensions["A"].width = 100

    ws["A1"] = "Notes & Usage Guide"
    ws["A1"].font = title_font()

    notes_lines = [
        ("As-of date", db["as_of"]),
        ("Schema version", db["schema_version"]),
        ("Source data", "sales/estimating/pricing-db.json"),
        ("Regenerate workbook", "python build_estimator_workbook.py"),
        ("", ""),
        ("How to use", ""),
        ("1.", "Open the Quote Input sheet. Fill yellow cells from top to bottom."),
        ("2.", "Select Project Type first. The engine adjusts what gets included/excluded based on the matrix in 'Project Type Behavior'."),
        ("3.", "The big sell price at the top of Quote Input updates live as you edit inputs."),
        ("4.", "For client proposals, copy the formatted block from the 'Client Range' sheet."),
        ("5.", "For scope language by project type, use 'Scope Builder' to copy the appropriate blocks."),
        ("", ""),
        ("Pricing updates", ""),
        ("To change a price", "Edit yellow cells on Pricing DB sheet, OR edit pricing-db.json and regenerate."),
        ("Quarterly review", "Re-pull Menards Wausau prices every quarter. Cedar is especially volatile."),
        ("", ""),
        ("Math reminders", ""),
        ("Margin vs markup", "Sell Price = Cost / (1 - margin). 20% margin = Cost / 0.80."),
        ("Multipliers", "Chained as a product. 1.22 height x 1.10 grade = 1.342 combined."),
        ("Project type matrix", "Y = always include; N = always exclude; OPT = include if user enters non-zero; 'Inspect' = use inspection allowance; 'Light' = use Height + Grade multipliers only."),
        ("", ""),
        ("Bug fixes from inherited tool", ""),
        ("E28-E31 fixed", "John's original workbook had E28/E29/E30/E31 (sell price + low/high range + per-SF) referencing the wrong source cell. CWDB v1 fixes these by computing from explicit SUBTOTAL."),
        ("Adder lookups fixed", "John's original used hard-coded strings ('Deck Lighting') in INDEX/MATCH formulas. CWDB v1 uses keyed lookups against an editable adder table."),
        ("Waste applied to decking only", "John applied waste % to combined framing+decking. CWDB v1 applies waste only to decking material (framing is precise; no waste)."),
    ]

    r = 3
    for label, content in notes_lines:
        ws.cell(row=r, column=1).value = (f"{label}: {content}" if label and content
                                          else (label or content))
        ws.cell(row=r, column=1).font = (label_font(bold=True) if label and not content
                                          else label_font())
        ws.cell(row=r, column=1).alignment = WRAP_LEFT
        if content and len(content) > 80:
            ws.row_dimensions[r].height = 30
        r += 1

    return ws


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    db = load_pricing()
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    build_pricing_db_v2(wb, db)     # cursor-based, records PDB ranges
    build_project_behavior(wb, db)  # records PB ranges
    build_quote_input(wb, db)       # inserts at position 0
    build_materials_list(wb, db)
    build_estimate_summary(wb, db)
    build_client_range(wb, db)
    build_scope_builder(wb, db)
    build_notes(wb, db)

    # Tab colors
    color_map = {
        "Quote Input": ORANGE,
        "Materials List": SKY,
        "Pricing DB": SLATE,
        "Project Type Behavior": GREY,
        "Estimate Summary": ORANGE,
        "Client Range": SKY,
        "Scope Builder": GREY,
        "Notes": GREY,
    }
    for name, color in color_map.items():
        if name in wb.sheetnames:
            wb[name].sheet_properties.tabColor = color

    wb.save(OUTPUT_PATH)
    print(f"Workbook written: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
