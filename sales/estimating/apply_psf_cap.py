"""
Apply a per-square-foot price ceiling to a folder of already-generated CWDB
estimates and regenerate the PDF + workbook in place.

LEGACY (v1 engine era) - deprecated 2026-07-09
----------------------------------------------
The estimator moved to the v2 explicit-labor engine (pricing-db-v2.json):
prices are built from true cost, the app shows a real Breakeven rung with a
below-cost warning, and the Excel workbook this tool reads inputs from is no
longer attached to estimate emails (replaced by the Materials & Hardware List
PDF). The $70/$100 psf ceilings were a patch over v1's inflated sell rates;
under v2 a cap that back-solves margin below breakeven is a visible
money-loser. Keep this file only to regenerate PRE-CUTOVER estimates from
their workbooks; do not run it against v2 estimates.

Why this existed
----------------
The deck engine prices bottom-up (sell = subtotal / (1 - margin)) and small
decks carry the fixed allowances (mobilization, permit, dumpster) plus margin
over very little area, so the blended $/sq ft runs far above market. For early
jobs Jim wants a hard ceiling by decking class:

    TimberTech / composite decking   <= $100 / sq ft   (total project / deck SF)
    Pressure-treated / wood decking  <=  $70 / sq ft

This reads each estimate's saved inputs back out of its Quote Input workbook,
recomputes the engine, and for any per-sf job over its ceiling forces the total
to (cap x deck SF). It does that through the engine's own margin lever
(margin = 1 - subtotal / target) so build_line_items re-scales every customer
line proportionally and the document stays internally consistent. Client
contact details (phone, email) are recovered from the sibling PDF, which is the
faithful record of what was sent.

Fence jobs are priced by linear foot, not deck SF, so the $/sf ceilings do not
apply; they are reported and left untouched.

Usage
-----
    python apply_psf_cap.py "C:/Users/jslog/Downloads/estimates"

Originals should be backed up by the caller; this overwrites in place.
"""

from __future__ import annotations

import re
import sys
import tempfile
from pathlib import Path

import fitz  # PyMuPDF
from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent          # sales/estimating
ESTIMATES_DIR = SCRIPT_DIR.parent / "estimates"        # sales/estimates
APP_DIR = SCRIPT_DIR / "streamlit_app"
for p in (SCRIPT_DIR, ESTIMATES_DIR, APP_DIR):
    if str(p) not in sys.path:
        sys.path.insert(0, str(p))

from deck_calculator import (  # noqa: E402
    load_pricing,
    compute_engine,
    find_project_type,
    build_estimate_json,
)
from generate_estimate_pdf import generate_pdf  # noqa: E402
from build_estimator_workbook import QI  # noqa: E402
from populate_workbook import INPUT_CELL_MAP, populate as populate_workbook  # noqa: E402

CWDB_PHONE_DIGITS = "7155447941"
CWDB_EMAIL = "info@cwdeckbuilders.com"

NUMERIC_KEYS = (
    "length", "depth", "fascia_lf", "railing_lf", "stair_runs", "stair_treads",
    "stair_landings", "stain_sf", "board_repairs", "joist_repair_lf",
    "skirting_sf", "lighting_fix", "bench_count", "privacy_screen_lf",
    "fence_lf", "walk_gates", "drive_gates", "tearout_lf",
)

DEFAULTS = dict(
    project_type=None, length=1, depth=1, decking_material=None, decking_color=None,
    railing_material=None, railing_color=None, framing_material=None,
    height="1 Story / Low", grade="Flat / Normal Access",
    complexity="Simple Rectangle", market="Normal Schedule",
    border_style="Pencil Border", fascia_lf=0, railing_lf=0, stair_runs=0,
    stair_treads=0, stair_landings=0, wraparound="No", stain_sf=0,
    stain_type="Solid / Paint-and-Sealer", stain_coats=1, stain_color=None,
    board_repairs=0, joist_repair_lf=0, hardware_inc="No", skirting_sf=0,
    lighting_fix=0, bench_count=0, privacy_screen_lf=0, hot_tub="No",
    fence_material=None, fence_height="6", fence_color=None, fence_lf=0,
    walk_gates=0, drive_gates=0, tearout_lf=0,
)


def cap_for_decking(name: str):
    """Return (class_label, $/sf ceiling) for a decking material, or (label, None)
    when no ceiling is specified for that class."""
    n = (name or "").lower()
    if "pressure-treated" in n or "pressure treated" in n or "pt " in n or "pine" in n:
        return ("Pressure-treated", 70)
    if "timbertech" in n or "trex" in n or "composite" in n or "azek" in n:
        return ("TimberTech / composite", 100)
    if "cedar" in n:
        return ("Cedar", None)   # not specified by owner
    return (name or "?", None)


def read_inputs_from_xlsx(xlsx_path: Path) -> dict:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["Quote Input"]
    inp = dict(DEFAULTS)
    for key, cell in INPUT_CELL_MAP.items():
        v = ws[cell].value
        if v is not None:
            inp[key] = v
    for k in NUMERIC_KEYS:
        try:
            inp[k] = int(inp[k])
        except (TypeError, ValueError):
            pass
    try:
        inp["stain_coats"] = int(inp.get("stain_coats") or 1)
    except (TypeError, ValueError):
        inp["stain_coats"] = 1
    return inp


def recover_contact_from_pdf(pdf_path: Path) -> tuple[str, str]:
    """Return (phone, email) for the homeowner, read from the sent PDF."""
    if not pdf_path.exists():
        return ("", "")
    doc = fitz.open(str(pdf_path))
    text = doc[0].get_text("text")
    emails = re.findall(r"[\w.+-]+@[\w.-]+\.\w+", text)
    email = next((e for e in emails if e.lower() != CWDB_EMAIL), "")
    phones = re.findall(r"\(?\d{3}\)?[ .-]?\d{3}[ .-]?\d{4}", text)
    phone = next(
        (p for p in phones if re.sub(r"\D", "", p) != CWDB_PHONE_DIGITS), ""
    )
    return (phone, email)


def extract_renderings(src_pdf: Path, out_dir: Path) -> list[dict]:
    """Pull the AI mock-up images off the Design Mock-Up page of a previously
    generated PDF, in left-to-right order, and save them as PNGs. The
    'AI SIMULATION' watermark is burned into the image, so extracting the
    embedded bitmap preserves it. Returns [{'path', 'caption'}] for re-injection.
    """
    if not src_pdf.exists():
        return []
    doc = fitz.open(str(src_pdf))
    mock_page = None
    for pg in doc:
        if "Design Mock-Up" in pg.get_text("text"):
            mock_page = pg
            break
    if mock_page is None:
        return []

    placed = []
    for img in mock_page.get_images(full=True):
        xref = img[0]
        rects = mock_page.get_image_rects(xref)
        if not rects:
            continue
        placed.append((min(r.x0 for r in rects), xref))
    placed.sort(key=lambda t: t[0])  # left to right

    out_dir.mkdir(parents=True, exist_ok=True)
    renderings = []
    for i, (_, xref) in enumerate(placed):
        pix = fitz.Pixmap(doc, xref)
        if pix.n - pix.alpha >= 4:           # CMYK -> RGB
            pix = fitz.Pixmap(fitz.csRGB, pix)
        dest = out_dir / f"{src_pdf.stem}-mockup-{i}.png"
        pix.save(str(dest))
        renderings.append({"path": str(dest), "caption": ""})
    return renderings


def recover_schedule_start(pdf_path: Path) -> str | None:
    if not pdf_path.exists():
        return None
    doc = fitz.open(str(pdf_path))
    text = "\n".join(p.get_text("text") for p in doc)
    m = re.search(r"Targeted start:\s*(.+)", text)
    return m.group(1).strip() if m else None


def regenerate(xlsx_path: Path, db, dry_run: bool = False) -> dict:
    stem = xlsx_path.stem
    pdf_path = xlsx_path.with_suffix(".pdf")
    # The backup original is the faithful record (still carries mock-up images
    # and the as-sent client block); prefer it as the source of truth.
    backup_pdf = xlsx_path.parent / "_original-pre-cap" / f"{stem}.pdf"
    source_pdf = backup_pdf if backup_pdf.exists() else pdf_path
    inputs = read_inputs_from_xlsx(xlsx_path)

    name = inputs.pop("client_name", None) or "Homeowner"
    address = inputs.pop("address", None) or "Central Wisconsin"
    phone, email = recover_contact_from_pdf(source_pdf)
    client_info = {
        "name": name,
        "address_line": address,
        "phone": phone or "(715) 544-7941",
        "email": email or CWDB_EMAIL,
    }

    pt = find_project_type(db, inputs["project_type"])
    is_fence = pt["matrix"].get("fence") == "Y"

    eng = compute_engine(db, inputs)
    sell = eng["sell_price"]
    subtotal = eng["subtotal_cost"]

    result = {
        "stem": stem, "name": name, "project_type": inputs["project_type"],
        "is_fence": is_fence, "old_sell": round(sell),
    }

    if is_fence:
        flf = eng.get("fence_lf", 0)
        result.update(unit=round(sell / flf, 2) if flf else 0, basis="LF",
                      action="left unchanged (per-sf cap N/A to fence)")
        return result

    sf = eng["deck_sf"]
    cls, cap = cap_for_decking(inputs.get("decking_material"))
    old_psf = sell / sf if sf else 0
    result.update(sf=sf, decking=inputs.get("decking_material"),
                  klass=cls, cap=cap, old_psf=round(old_psf, 2), basis="SF")

    if cap is None:
        result["action"] = f"left unchanged (no ceiling defined for {cls})"
        return result
    if old_psf <= cap + 1e-9:
        result.update(new_sell=round(sell), new_psf=round(old_psf, 2),
                      action=f"already <= ${cap}/sf, unchanged")
        return result

    target = cap * sf
    margin_override = 1 - subtotal / target   # engine lever; hits target exactly
    capped_inputs = dict(inputs)
    capped_inputs["margin"] = margin_override

    result.update(new_sell=round(target), new_psf=round(target / sf, 2),
                  margin_override=round(margin_override, 4),
                  action=f"CAPPED to ${cap}/sf")

    if dry_run:
        return result

    # --- Regenerate the PDF (homeowner deliverable) ---
    with tempfile.TemporaryDirectory() as tmp:
        estimate = build_estimate_json(capped_inputs, client_info, db)
        start = recover_schedule_start(source_pdf)
        if start:
            estimate["schedule"]["start"] = start
            estimate["schedule"]["start_label"] = "Targeted start"
        renderings = extract_renderings(source_pdf, Path(tmp))
        if renderings:
            estimate["renderings"] = renderings
        result["mockups"] = len(renderings)
        generate_pdf(estimate, pdf_path)

        # Verify the rendered total equals the target (within rounding)
        rendered = sum(it[1] for it in estimate["line_items"])
        result["rendered_total"] = rendered
        result["render_ok"] = abs(rendered - round(target)) <= 2

    # --- Regenerate the workbook so Jim's internal record matches ---
    populate_workbook(capped_inputs, client_info, xlsx_path)
    wb = load_workbook(xlsx_path)
    wb["Quote Input"][QI.TARGET_MARGIN] = margin_override
    wb.save(xlsx_path)

    return result


def main():
    folder = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(
        "C:/Users/jslog/Downloads/estimates")
    dry = "--dry-run" in sys.argv
    db = load_pricing()
    xlsxs = sorted(p for p in folder.glob("*.xlsx") if not p.name.startswith("~"))
    print(f"{'JOB':30s} {'BASIS':5s} {'OLD':>10s} {'$/u':>8s} {'NEW':>10s} {'$/u':>7s}  ACTION")
    for x in xlsxs:
        r = regenerate(x, db, dry_run=dry)
        old = f"${r['old_sell']:,}"
        if r["is_fence"]:
            print(f"{r['stem']:30s} {'LF':5s} {old:>10s} {r['unit']:>8} {'-':>10s} {'-':>7s}  {r['action']}")
        else:
            oldu = f"${r['old_psf']:.0f}"
            new = f"${r.get('new_sell', r['old_sell']):,}"
            newu = f"${r.get('new_psf', r['old_psf']):.0f}"
            flag = "" if r.get("render_ok", True) else "  !! RENDER MISMATCH"
            mk = f" [{r['mockups']} mockup img]" if r.get("mockups") else ""
            print(f"{r['stem']:30s} {'SF':5s} {old:>10s} {oldu:>8s} {new:>10s} {newu:>7s}  {r['action']}{mk}{flag}")


if __name__ == "__main__":
    main()
