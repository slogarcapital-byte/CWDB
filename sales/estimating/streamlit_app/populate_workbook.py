"""
Populate a copy of CWDB_Deck_Estimator_v1.xlsx with web-app inputs.

Opens the template workbook, writes user-entered values into the Quote Input
sheet's input cells, and saves a per-quote copy. All formulas in the template
remain intact and Excel recalculates them on open.

The cell-address map (`QI`) is imported from build_estimator_workbook so we
have a single source of truth; if Quote Input cells move, the workbook
generator updates QI and this stays in sync.
"""

from __future__ import annotations

import sys
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
ESTIMATING_DIR = SCRIPT_DIR.parent
TEMPLATE_PATH = ESTIMATING_DIR / "CWDB_Deck_Estimator_v1.xlsx"

if str(ESTIMATING_DIR) not in sys.path:
    sys.path.insert(0, str(ESTIMATING_DIR))

from build_estimator_workbook import QI  # noqa: E402


INPUT_CELL_MAP: dict[str, str] = {
    # Client / project info
    "client_name": QI.CLIENT_NAME,
    "address": QI.ADDRESS,
    "estimator": QI.ESTIMATOR,
    "quote_date": QI.QUOTE_DATE,
    # Project type + dimensions
    "project_type": QI.PROJECT_TYPE,
    "length": QI.LENGTH,
    "depth": QI.DEPTH,
    # Materials
    "decking_material": QI.DECKING_MAT,
    "railing_material": QI.RAILING_MAT,
    "framing_material": QI.FRAMING_MAT,
    # Site conditions
    "height": QI.HEIGHT_COND,
    "grade": QI.GRADE_COND,
    "complexity": QI.COMPLEXITY_COND,
    "market": QI.MARKET_COND,
    # Scope detail
    "border_style": QI.BORDER_STYLE,
    "railing_lf": QI.RAILING_LF,
    "fascia_lf": QI.FASCIA_LF,
    "stair_runs": QI.STAIR_RUNS,
    "stair_treads": QI.STAIR_TREADS,
    "stair_landings": QI.STAIR_LANDINGS,
    "wraparound": QI.WRAPAROUND,
    # Stain
    "stain_sf": QI.STAIN_SF,
    "stain_type": QI.STAIN_TYPE,
    "stain_coats": QI.STAIN_COATS,
    # Repair bucket
    "board_repairs": QI.BOARD_REPAIRS,
    "joist_repair_lf": QI.JOIST_REPAIR_LF,
    "hardware_inc": QI.HARDWARE_INC,
    # Adders
    "skirting_sf": QI.SKIRTING_SF,
    "lighting_fix": QI.LIGHTING_FIX,
    "bench_count": QI.BENCH_COUNT,
    "privacy_screen_lf": QI.PRIV_SCREEN_LF,
    "hot_tub": QI.HOT_TUB,
}


def populate(
    inputs: dict[str, Any],
    client_info: dict[str, str],
    output_path: Path,
    template_path: Path = TEMPLATE_PATH,
) -> Path:
    """Open the template, write inputs + client info to Quote Input cells,
    save to output_path. Returns output_path."""
    if not template_path.exists():
        raise FileNotFoundError(f"Template workbook not found: {template_path}")

    wb = load_workbook(template_path)
    ws = wb["Quote Input"]

    flat = {
        "client_name": client_info.get("name", ""),
        "address": client_info.get("address_line", ""),
        "estimator": "James Slogar",
        "quote_date": date.today().isoformat(),
        **{k: v for k, v in inputs.items() if k in INPUT_CELL_MAP},
    }

    for key, value in flat.items():
        cell_ref = INPUT_CELL_MAP.get(key)
        if cell_ref is None or value is None:
            continue
        ws[cell_ref] = value

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
